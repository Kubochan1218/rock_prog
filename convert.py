import openpyxl
import customtkinter as ctk
from tkinter import filedialog, messagebox


# アプリ全体のテーマカラー設定
ctk.set_appearance_mode("System")  # デフォルトは"System"
ctk.set_default_color_theme("green")  # "blue", "green", "dark-blue"
FONT_NAME = 'Yu Gothic UI'

class ExcelConverter:
    def __init__(self, master):
        self.master = master
        master.title('Excel出席データ変換ツール')
        master.geometry('540x420')
        master.minsize(540, 420)
        master.iconbitmap(default='convert_icon.ico')  # アイコン設定（Windows用）

        self.show_top_view()

    def clear_widgets(self):
        """画面のウィジェットをクリアする"""
        for widget in self.master.winfo_children():
            widget.destroy()

    def show_top_view(self):
        """画面表示"""
        self.clear_widgets()

        self.top_view = ctk.CTkFrame(self.master, fg_color="transparent")
        self.top_view.pack(fill='both', expand=True, padx=20)

        # ファイル選択ボタン
        title_label = ctk.CTkLabel(self.top_view, text='出席データ形式変換', font=(FONT_NAME, 20, 'bold'))
        title_label.pack(pady=(15, 5), anchor='w')
        file_label = ctk.CTkLabel(self.top_view, text='出席データ形式を変換するExcelファイルを選択してください。', font=(FONT_NAME, 14), text_color='gray50')
        file_label.pack(pady=10, anchor='w')

        file_path_var = ctk.StringVar(value='Excelファイルを選択してください')
        file_frame = ctk.CTkFrame(self.top_view, fg_color="transparent")
        file_frame.pack(anchor='w', fill='x')
            
        file_entry = ctk.CTkEntry(file_frame, textvariable=file_path_var, width=350, font=(FONT_NAME, 16), state='disabled')
        file_entry.pack(side='left', padx=(0, 10))

        def select_file():
            f_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
            if f_path:
                file_path_var.set(f_path)
                self.file_path = f_path

        btn_file = ctk.CTkButton(file_frame, text='ファイルを選択', width=120, fg_color='#80d4ff', text_color='black', font=(FONT_NAME, 16), command=select_file)
        btn_file.pack(side='left')

        self.copy_check_box = ctk.CTkCheckBox(self.top_view, text='変換前に元のファイルをコピーする', font=(FONT_NAME, 14), text_color='gray50')
        self.copy_check_box.pack(pady=10, anchor='w')

        # 変換ボタン
        btn_convert = ctk.CTkButton(self.top_view, text='変換開始', width=200, height=40, fg_color="#20e030", text_color='black', font=(FONT_NAME, 16, 'bold'), command=self.check_status)
        btn_convert.pack(side='bottom', pady=20)

    
    def check_status(self):
        if not hasattr(self, 'file_path') or not self.file_path:
            messagebox.showwarning("警告", "変換するExcelファイルを選択してください。")
            return
        try:
            self.copy_file = self.copy_check_box.get()
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            ws = wb['出欠状況']
            self.convert_dict = {
                '出席': '出席',
                '連絡あり': '連絡あり',
                '無断欠席': '無断欠席',
                'オ': 'オ',
                '忌引': '忌引',
                '○': '出席',
                '〇': '出席',
                '△': '連絡あり',
                '×': '無断欠席',
                '': '',
                None: None
            }
            additional_status = []
            for col in range(7, ws.max_column + 1):
                cell_value = ws.cell(row=2, column=col).value
                if cell_value is not None:
                    for row in range(3, ws.max_row + 1):
                        status = ws.cell(row=row, column=col).value
                        if status not in self.convert_dict and status not in additional_status:
                            additional_status.append(status)
            
            if additional_status:
                self.set_additional_status(additional_status)
                return
            else:
                self.convert()
        except Exception as e:
            messagebox.showerror("エラー", f"Excelファイルの変換中にエラーが発生しました: {e}")
    
    def set_additional_status(self, additional_status):
        """追加の出席状況を設定する"""
        self.clear_widgets()

        self.top_view = ctk.CTkFrame(self.master, fg_color="transparent")
        self.top_view.pack(fill='both', expand=True, padx=20)

        title_label = ctk.CTkLabel(self.top_view, text='変換内容の登録', font=(FONT_NAME, 20, 'bold'))
        title_label.pack(pady=(15, 5), anchor='w')
        file_label = ctk.CTkLabel(self.top_view, text='変換後の出席状態を選択してください。', font=(FONT_NAME, 14), text_color='gray50')
        file_label.pack(pady=10, anchor='w')

        # 変換前と変換後の出席状態を表示するヘッダー
        header_frame = ctk.CTkFrame(self.top_view, fg_color="transparent")
        header_frame.pack(anchor='center', fill='x')
        header_frame.grid_columnconfigure(0, weight=1, uniform="col0")
        header_frame.grid_columnconfigure(1, weight=1, uniform="col0")
        header_before_label = ctk.CTkLabel(header_frame, text='変換前の出席状態', font=(FONT_NAME, 14, 'bold'))
        header_before_label.grid(row=0, column=0, sticky='ew', padx=5)
        header_after_label = ctk.CTkLabel(header_frame, text='変換後の出席状態', font=(FONT_NAME, 14, 'bold'))
        header_after_label.grid(row=0, column=1, sticky='ew', padx=5)

        convert_scroll_frame = ctk.CTkScrollableFrame(self.top_view)
        convert_scroll_frame.pack(pady=5, anchor='center', fill='both', expand=True)
        for status in additional_status:
            row_frame = ctk.CTkFrame(convert_scroll_frame, fg_color="transparent")
            row_frame.pack(anchor='center', fill='x', pady=5)
            row_frame.grid_columnconfigure(0, weight=1, uniform="col0")
            row_frame.grid_columnconfigure(1, weight=1, uniform="col0")

            status_label = ctk.CTkLabel(row_frame, text=status, font=(FONT_NAME, 14))
            status_label.grid(row=0, column=0, sticky='ew', padx=5)

            status_var = ctk.StringVar(value='出席')
            status_option_menu = ctk.CTkOptionMenu(row_frame, variable=status_var, values=['出席', '連絡あり', '無断欠席', ''], font=(FONT_NAME, 14), fg_color=('gray70', 'gray30'), button_color=('gray80', 'gray20'), button_hover_color=('gray90', 'gray25'))
            status_option_menu.grid(row=0, column=1, sticky='ew', padx=5)

            self.convert_dict[status] = status_var
        
        convert_button = ctk.CTkButton(self.top_view, text='変換開始', width=200, height=40, fg_color="#20e030", text_color='black', font=(FONT_NAME, 16, 'bold'), command=self.convert)
        convert_button.pack(side='bottom', pady=(5, 20))

    def convert(self):
        """Excelファイルの変換処理"""
        try:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb['出欠状況']

            # 変換前の出席状態を変換後の出席状態に置き換える
            for col in range(7, ws.max_column + 1):
                for row in range(3, ws.max_row + 1):
                    status = ws.cell(row=row, column=col).value
                    if status in self.convert_dict:
                        new_status = self.convert_dict[status]
                        if isinstance(new_status, ctk.StringVar):
                            new_status = new_status.get()
                        ws.cell(row=row, column=col).value = new_status

            # 元のファイルをコピーする場合
            if self.copy_file == 1:
                copy_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
                if copy_path:
                    wb.save(copy_path)
            else:
                wb.save(self.file_path)
            messagebox.showinfo("完了", "Excelファイルの変換が完了しました。")
        except Exception as e:
            self.show_top_view()
            messagebox.showerror("エラー", f"Excelファイルの変換中にエラーが発生しました: {e}")


if __name__ == '__main__':
    root = ctk.CTk()            
    app = ExcelConverter(root)
    root.mainloop()