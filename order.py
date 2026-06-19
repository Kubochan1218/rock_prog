import csv
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from datetime import datetime, timedelta
from models import LiveSchedule
from models import BandInfo
import json
import os
import sys
import shutil

"""
「csvファイルが見つかりません。」
↓現在の作業ディレクトリを確認して、パスを指定しなおす。
import os
print("現在の作業ディレクトリ：", os.getcwd())
"""

#バンド情報が格納されたCSVファイルを読み込む
def load_band_infos_from_csv(csv_path):
    band_list = []

    with open(csv_path, newline='', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            name = row["バンド名"]
            minutes = int(row["演奏時間"])
            dates_str = row["出演日"]

            # 出演日をリストに変換（複数日ある場合）
            date_list = []
            for part in dates_str.split(";"):
                try:
                    date = datetime.strptime(part.strip(), "%Y-%m-%d").date()
                    date_list.append(date)
                except ValueError:
                    pass  # 不正な日付形式はスキップ

            band = BandInfo(name, minutes, date_list)
            band_list.append(band)

    return band_list


class OrderWindow(tk.Toplevel):
    def __init__(self, master, schedules):
        super().__init__(master)
        self.title("出演順設定")
        self.geometry("600x500")
        self.attributes("-topmost", True)
        self.schedules = schedules  # List of LiveSchedule
        self.band_data = load_band_infos_from_csv("./kadai5/bands.csv")

        # --- メニューボタン ---
        menu_frame = tk.Frame(self)
        menu_frame.pack(fill=tk.X, padx=10, pady=5)
        btn_save = tk.Button(menu_frame, text="保存", command=self.save_schedule, font=("Meiryo", 12), width=7, height=1)
        btn_save.pack(side=tk.LEFT, padx=10)
        btn_load = tk.Button(menu_frame, text="読み込み", command=self.load_schedule, font=("Meiryo", 12), width=7, height=1)
        btn_load.pack(side=tk.LEFT, padx=10)
        btn_export_pdf = tk.Button(menu_frame, text="PDF出力", command=self.export_pdf, font=("Meiryo", 12), width=8, height=1, bg="#e1bee7")
        btn_export_pdf.pack(side=tk.LEFT, padx=10)
        btn_export_excel = tk.Button(menu_frame, text="Excel出力", command=self.export_excel, font=("Meiryo", 12), width=10, height=1, bg="#b2dfdb")
        btn_export_excel.pack(side=tk.LEFT, padx=10)

        self.tab_control = ttk.Notebook(self)
        self.tab_control.pack(expand=True, fill="both")

        self.tabs = {}  # tab: { "combo": ComboBox, "frame": Frame }
        
        self.used_band_names = set() #追加済みバンド

        self.create_tabs()
        self._last_font_copied = False

    def _ensure_font_file(self, filename, friendly_name="フォント"):
        """アプリフォルダに filename がなければ Windows のフォントフォルダからコピーを試みる。
        成功した場合はコピー先のパスを返す。失敗した場合は None を返す（かつエラーメッセージを表示）。
        """
        app_dir = os.path.dirname(os.path.abspath(__file__))
        target_path = os.path.join(app_dir, filename)
        if os.path.exists(target_path):
            return target_path

        windir = os.environ.get("WINDIR", "C:\\Windows")
        fonts_dir = os.path.join(windir, "Fonts")
        candidates = []
        try:
            for f in os.listdir(fonts_dir):
                if f.lower() == filename.lower() or filename.split('.')[0].lower() in f.lower():
                    candidates.append(os.path.join(fonts_dir, f))
        except Exception:
            candidates = []

        if not candidates:
            messagebox.showerror(f"{friendly_name} コピー失敗", f"{filename} が見つかりません。\nシステムフォントフォルダに {filename} が存在しないようです。")
            return None

        for src in candidates:
            try:
                shutil.copy2(src, target_path)
                self._last_font_copied = True
                return target_path
            except Exception:
                continue

        messagebox.showerror(f"{friendly_name} コピー失敗", f"{filename} のコピーに失敗しました。手動で {filename} をアプリフォルダに置いてください。")
        return None


    def create_tabs(self):
        for sched in self.schedules:
            tab = ttk.Frame(self.tab_control)
            self.tab_control.add(tab, text=sched.date.strftime("%m/%d"))

            # --- 上部：バンド追加パネル ---
            top_frame = tk.Frame(tab)
            top_frame.pack(fill=tk.X, padx=10, pady=5)

            label = tk.Label(top_frame, text="バンドを選択：")
            label.pack(side=tk.LEFT)

            combo = ttk.Combobox(top_frame, width=30, state="readonly")
            combo.pack(side=tk.LEFT, padx=5)

            btn_add = tk.Button(top_frame, text="追加", command=lambda c=combo, t=tab: self.add_band(c, t), font=("Meiryo", 12, "bold"), width=8, height=1, bg="#e0f7fa")
            btn_add.pack(side=tk.LEFT, padx=8)
            btn_special = tk.Button(top_frame, text="特別枠追加", command=lambda t=tab: self.add_special_frame(t), font=("Meiryo", 12), width=12, height=1, bg="#ffe0b2")
            btn_special.pack(side=tk.LEFT, padx=8)

            # --- バンド表示エリア（スクロール対応） ---
            canvas = tk.Canvas(tab, borderwidth=0, background="#f8f8f8")
            frame_container = tk.Frame(canvas, background="#f8f8f8")
            vsb = tk.Scrollbar(tab, orient="vertical", command=canvas.yview)
            canvas.configure(yscrollcommand=vsb.set)
            vsb.pack(side="right", fill="y")
            canvas.pack(side="left", fill="both", expand=True)
            canvas.create_window((0, 0), window=frame_container, anchor="nw")
            def on_frame_configure(event, c=canvas, f=frame_container):
                c.configure(scrollregion=c.bbox("all"))
            frame_container.bind("<Configure>", on_frame_configure)

            # --- マウスホイールでスクロール ---
            def _on_mousewheel(event, c=canvas):
                if event.delta:
                    c.yview_scroll(int(-1*(event.delta/120)), "units")
                elif event.num == 4:  # Linux
                    c.yview_scroll(-1, "units")
                elif event.num == 5:
                    c.yview_scroll(1, "units")
            canvas.bind_all("<MouseWheel>", _on_mousewheel)
            canvas.bind_all("<Button-4>", _on_mousewheel)
            canvas.bind_all("<Button-5>", _on_mousewheel)

            # --- 出演可能バンドだけを ComboBox に追加 ---
            available_bands = [
                b for b in self.band_data if sched.date in b.available_dates
            ]
            combo["values"] = [b.name for b in available_bands]

            # 各タブの情報を記録
            self.tabs[tab] = {
                "combo": combo,
                "frame": frame_container,  # スクロール内のフレーム
                "bands": [],  # 追加されたバンドや特別枠
                "band_objs": available_bands,
                "canvas": canvas,
                "scrollbar": vsb
            }

    #バンドを追加
    def add_band(self, combo, tab):
        band_name = combo.get()
        if not band_name:
            return
        tab_info = self.tabs[tab]
        # 重複チェック
        if any(b['type']=='band' and b['name']==band_name for b in tab_info["bands"]):
            messagebox.showwarning("重複追加", f"{band_name} はすでに他の日程で追加されています。")
            return
        # BandInfoオブジェクトを取得
        band = self.find_band_by_name(band_name, tab_info["band_objs"])
        if not band:
            messagebox.showerror("エラー", "バンド情報が見つかりません")
            return

        # バンド名追加・ComboBox更新
        tab_info["bands"].append({'type': 'band', 'name': band_name, 'minutes': band.performance_minutes})
        self.used_band_names.add(band_name)
        for t in self.tabs.keys():
            self.refresh_combo(t)

        # バンド表示エリアを再描画
        self.update_band_frames(tab)

    def add_special_frame(self, tab):
        # ダイアログで種別・分数・リハの場合はバンド名を選択
        def on_ok():
            kind = var_kind.get()
            try:
                minutes = int(entry_min.get())
            except ValueError:
                messagebox.showerror("エラー", "分数は整数で入力してください")
                return
            if kind == 'リハ':
                band_name = combo_band.get().strip()
                if not band_name:
                    messagebox.showerror("エラー", "バンド名を選択してください")
                    return
                item = {'type': 'rehearsal', 'name': band_name, 'minutes': minutes}
            elif kind == '休憩':
                item = {'type': 'break', 'minutes': minutes}
            elif kind == '転換':
                item = {'type': 'change', 'minutes': minutes}
            else:
                return
            self.tabs[tab]["bands"].append(item)
            win.destroy()
            self.update_band_frames(tab)
        win = tk.Toplevel(self)
        win.title("特別枠追加")
        win.geometry("250x120")  # 横幅を250に修正
        tk.Label(win, text="種別").grid(row=0, column=0)
        var_kind = tk.StringVar(value='休憩')
        kinds = ['休憩', '転換', 'リハ']
        combo_kind = ttk.Combobox(win, values=kinds, textvariable=var_kind, state="readonly", width=8, justify='center')
        combo_kind.grid(row=0, column=1, padx=5, pady=2)
        combo_kind.configure(justify='center')
        tk.Label(win, text="分数").grid(row=1, column=0)
        entry_min = tk.Entry(win, width=8, justify='center')
        entry_min.grid(row=1, column=1, padx=5, pady=2)
        tk.Label(win, text="バンド名(リハのみ)").grid(row=2, column=0)
        # バンド名コンボボックス
        all_band_names = [b.name for b in self.tabs[tab]['band_objs']]
        combo_band = ttk.Combobox(win, values=all_band_names, state="readonly", width=18, justify='center')
        combo_band.grid(row=2, column=1, padx=5, pady=2)
        combo_band.configure(justify='center')
        btn = tk.Button(win, text="OK", command=on_ok)
        btn.grid(row=3, column=0, columnspan=2, pady=5)
        def on_kind_change(event):
            if var_kind.get() == 'リハ':
                combo_band.config(state='readonly')
            else:
                combo_band.set("")
                combo_band.config(state='disabled')
        combo_kind.bind('<<ComboboxSelected>>', on_kind_change)
        on_kind_change(None)

    # バンド表示エリアを枠順＋時刻付きで再描画
    def update_band_frames(self, tab):
        tab_info = self.tabs[tab]
        frame = tab_info["frame"]
        # 既存の表示をクリア
        for widget in frame.winfo_children():
            widget.destroy()

        # スケジュール情報取得
        sched = None
        for s in self.schedules:
            if s.date.strftime("%m/%d") == self.tab_control.tab(tab, "text"):
                sched = s
                break
        if not sched:
            return

        # 終了時刻
        end_time_limit = datetime.strptime(sched.end_time, "%H:%M")
        # 開始時刻
        current_time = datetime.strptime(sched.start_time, "%H:%M")
        for idx, item in enumerate(tab_info["bands"]):
            if item['type'] == 'band':
                name = item['name']
                minutes = item['minutes']
                label_text = f"{name}（{minutes}分）"
            elif item['type'] == 'break':
                name = None
                minutes = item['minutes']
                label_text = f"休憩（{minutes}分）"
            elif item['type'] == 'change':
                name = None
                minutes = item['minutes']
                label_text = f"転換（{minutes}分）"
            elif item['type'] == 'rehearsal':
                name = item['name']
                minutes = item['minutes']
                label_text = f"リハ（{name}）（{minutes}分）"
            else:
                continue
            start_str = current_time.strftime("%H:%M")
            end_time = current_time + timedelta(minutes=minutes)
            end_str = end_time.strftime("%H:%M")

            wrapper = tk.Frame(frame, bd=1, relief="solid", padx=5, pady=5, width=560, height=36)
            wrapper.pack(fill=tk.X, expand=True, pady=2)
            wrapper.pack_propagate(False)  # サイズ固定

            label = tk.Label(wrapper, text=f"{label_text} {start_str}～{end_str}", anchor="w", font=("Meiryo", 12, "bold"))
            label.pack(side=tk.LEFT, fill=tk.X, expand=True)

            btn_del = tk.Button(wrapper, text="×", width=3, fg="red", font=("Meiryo", 12))
            btn_del.pack(side=tk.RIGHT)

            btn_down = tk.Button(wrapper, text="↓", width=2, font=("Meiryo", 12))
            btn_down.pack(side=tk.RIGHT, padx=2)
            btn_up = tk.Button(wrapper, text="↑", width=2, font=("Meiryo", 12))
            btn_up.pack(side=tk.RIGHT, padx=2)

            def remove_item(idx=idx):
                removed = tab_info["bands"][idx]
                del tab_info["bands"][idx]
                # バンド枠ならused_band_namesからも削除
                if removed.get("type") == "band":
                    self.used_band_names.discard(removed["name"])
                    for t in self.tabs:
                        self.refresh_combo(t)
                self.update_band_frames(tab)

            btn_up.config(command=lambda idx=idx: self.move_band(tab, idx, -1))
            btn_down.config(command=lambda idx=idx: self.move_band(tab, idx, 1))
            btn_del.config(command=remove_item)

            # 終了時刻超過警告を枠内に表示
            if end_time > end_time_limit:
                warn = tk.Label(wrapper, text="終了時刻を超過しています", fg="orange", font=("Meiryo", 10, "bold"))
                warn.pack(side=tk.BOTTOM, fill=tk.X, pady=2)

            current_time = end_time

    # 並べ替え
    def move_band(self, tab, idx, direction):
        tab_info = self.tabs[tab]
        new_idx = idx + direction
        if 0 <= new_idx < len(tab_info["bands"]):
            tab_info["bands"][idx], tab_info["bands"][new_idx] = tab_info["bands"][new_idx], tab_info["bands"][idx]
            self.update_band_frames(tab)

    #バンド削除時の処理
    def refresh_combo(self, tab):
        tab_info = self.tabs[tab]
        used_names = self.used_band_names

        filtered = [
            b.name for b in tab_info["band_objs"]
            if b.name not in used_names
        ]

        tab_info["combo"]["values"] = filtered

        if filtered:
            tab_info["combo"].current(0)
        else:
            tab_info["combo"].set("")


    #BandInfoを名前から検索
    def find_band_by_name(self, name, band_list):
        for b in band_list:
            if b.name == name:
                return b
        return None

    def save_schedule(self):
        data = {
            "schedules": [
                {
                    "date": s.date.strftime("%Y-%m-%d"),
                    "start_time": s.start_time,
                    "end_time": s.end_time
                } for s in self.schedules
            ],
            "bands": {
                self.tab_control.tab(tab, "text"): tabinfo["bands"]
                for tab, tabinfo in self.tabs.items()
            }
        }
        try:
            json_path = os.path.join(os.path.dirname(sys.argv[0]), "schedule_data.json")
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            messagebox.showinfo("保存", "スケジュールを保存しました。")
        except Exception as e:
            messagebox.showerror("保存エラー", str(e))

    def load_schedule(self):
        try:
            json_path = os.path.join(os.path.dirname(sys.argv[0]), "schedule_data.json")
            with open(json_path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            messagebox.showerror("読み込みエラー", str(e))
            return
        # 日程情報の復元（今回は日付・時刻が一致するものに枠を割り当てる）
        date2tab = {self.tab_control.tab(tab, "text"): tab for tab in self.tabs}
        bands_data = data.get("bands", {})
        self.used_band_names.clear()
        for tab_label, bands in bands_data.items():
            tab = date2tab.get(tab_label)
            if tab:
                self.tabs[tab]["bands"] = bands
        # 全タブのバンド枠からused_band_namesを再構築
        for tab in self.tabs:
            for b in self.tabs[tab]["bands"]:
                if b.get("type") == "band":
                    self.used_band_names.add(b["name"])
        for tab in self.tabs:
            self.refresh_combo(tab)
            self.update_band_frames(tab)
        messagebox.showinfo("読み込み", "スケジュールを読み込みました。")

    def export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
        except ImportError:
            messagebox.showerror("Excel出力エラー", "openpyxlライブラリが必要です。\npip install openpyxl でインストールしてください。")
            return
        # Excel出力前に Meiryo フォントファイルが存在するか確認（なければシステムからコピーを試みる）
        font_file = os.path.join(os.path.dirname(__file__), "meiryo.ttc")
        if not os.path.exists(font_file):
            if not self._ensure_font_file("meiryo.ttc", "Meiryo"):
                return

        wb = openpyxl.Workbook()
        for tab, tabinfo in self.tabs.items():
            tab_title = self.tab_control.tab(tab, "text")
            safe_title = tab_title.replace("/", "-")  # /を-に置換
            ws = wb.create_sheet(title=safe_title)
            ws.append(["開始時刻", "～", "終了時刻", "枠の名前"])
            # ヘッダにメイリオフォントを設定
            header_font = Font(name="Meiryo", size=12, bold=True)
            for col in range(1, 5):
                ws.cell(row=ws.max_row, column=col).font = header_font
            for idx, item in enumerate(tabinfo["bands"]):
                # 時刻計算
                sched = None
                for s in self.schedules:
                    if s.date.strftime("%m/%d") == self.tab_control.tab(tab, "text"):
                        sched = s
                        break
                if not sched:
                    continue
                current_time = datetime.strptime(sched.start_time, "%H:%M")
                for i in range(idx):
                    current_time += timedelta(minutes=tabinfo["bands"][i]["minutes"])
                start_str = current_time.strftime("%H:%M")
                end_time = current_time + timedelta(minutes=item["minutes"])
                end_str = end_time.strftime("%H:%M")
                name = item.get("name") if item.get("type") != "break" and item.get("type") != "change" else None
                if item["type"] == "band":
                    row = [start_str, "～", end_str, name]
                    ws.append(row)
                    for col in range(1, 5):
                        ws.cell(row=ws.max_row, column=col).fill = PatternFill(fill_type=None)
                        ws.cell(row=ws.max_row, column=col).font = Font(name="Meiryo", size=11, color="000000")
                else:
                    # 特別枠
                    if item["type"] == "break":
                        disp_name = f"休憩（{item['minutes']}分）"
                    elif item["type"] == "change":
                        disp_name = f"転換（{item['minutes']}分）"
                    elif item["type"] == "rehearsal":
                        disp_name = f"リハ（{item['name']}）（{item['minutes']}分）"
                    else:
                        disp_name = ""
                    row = [start_str, "～", end_str, disp_name]
                    ws.append(row)
                    for col in range(1, 5):
                        ws.cell(row=ws.max_row, column=col).fill = PatternFill(start_color="888888", end_color="888888", fill_type="solid")
                        ws.cell(row=ws.max_row, column=col).font = Font(name="Meiryo", size=11, color="FFFFFF")
                for col in range(1, 5):
                    ws.cell(row=ws.max_row, column=col).alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions["A"].width = 10
            ws.column_dimensions["B"].width = 4
            ws.column_dimensions["C"].width = 10
            ws.column_dimensions["D"].width = 24
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        try:
            wb.save("タイムテーブル.xlsx")
            msg = "タイムテーブル.xlsx を出力しました。"
            if getattr(self, '_last_font_copied', False):
                msg += "\n※フォントファイルコピー済み"
            messagebox.showinfo("Excel出力", msg)
            self._last_font_copied = False
        except Exception as e:
            messagebox.showerror("Excel出力エラー", str(e))

    def export_pdf(self):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
            from reportlab.lib.colors import black, white, HexColor
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            import os
            # MSゴシックフォント登録（ttfファイルはkadai5フォルダに配置）
            font_path = os.path.join(os.path.dirname(__file__), "BIZ-UDGothicR.ttc")
            if not os.path.exists(font_path):
                # システムフォントからコピーを試みる
                if not self._ensure_font_file("BIZ-UDGothicR.ttc", "BIZ UDゴシック"):
                    return
            pdfmetrics.registerFont(TTFont("BIZUDGothicR", font_path))
        except ImportError:
            messagebox.showerror("PDF出力エラー", "reportlabライブラリが必要です。\npip install reportlab でインストールしてください。")
            return
        try:
            c = canvas.Canvas("タイムテーブル.pdf", pagesize=A4)
            width, height = A4
            y_start = height - 50
            for tab, tabinfo in self.tabs.items():
                c.setFont("BIZUDGothicR", 16)
                c.drawString(50, y_start, f"日程: {self.tab_control.tab(tab, 'text')}")
                y = y_start - 30
                c.setFont("BIZUDGothicR", 12)
                c.drawString(50, y, "開始時刻")
                c.drawString(120, y, "～")
                c.drawString(160, y, "終了時刻")
                c.drawString(250, y, "枠の名前")
                y -= 24
                for idx, item in enumerate(tabinfo["bands"]):
                    sched = None
                    for s in self.schedules:
                        if s.date.strftime("%m/%d") == self.tab_control.tab(tab, "text"):
                            sched = s
                            break
                    if not sched:
                        continue
                    current_time = datetime.strptime(sched.start_time, "%H:%M")
                    for i in range(idx):
                        current_time += timedelta(minutes=tabinfo["bands"][i]["minutes"])
                    start_str = current_time.strftime("%H:%M")
                    end_time = current_time + timedelta(minutes=item["minutes"])
                    end_str = end_time.strftime("%H:%M")
                    if item["type"] == "band":
                        c.setFillColor(white)
                        c.setStrokeColor(black)
                        c.rect(45, y-2, 340, 22, fill=1, stroke=1)
                        c.setFillColor(black)
                        c.drawString(50, y, start_str)
                        c.drawString(120, y, "～")
                        c.drawString(160, y, end_str)
                        c.drawString(250, y, item["name"])
                    else:
                        if item["type"] == "break":
                            disp_name = f"休憩（{item['minutes']}分）"
                        elif item["type"] == "change":
                            disp_name = f"転換（{item['minutes']}分）"
                        elif item["type"] == "rehearsal":
                            disp_name = f"リハ（{item['name']}）（{item['minutes']}分）"
                        else:
                            disp_name = ""
                        c.setFillColor(HexColor("#888888"))
                        c.setStrokeColor(black)
                        c.rect(45, y-2, 340, 22, fill=1, stroke=1)
                        c.setFillColor(white)
                        c.drawString(50, y, start_str)
                        c.drawString(120, y, "～")
                        c.drawString(160, y, end_str)
                        c.drawString(250, y, disp_name)
                    y -= 24
                y_start = y - 40
                c.showPage()
            c.save()
            msg = "タイムテーブル.pdf を出力しました。"
            if getattr(self, '_last_font_copied', False):
                msg += "\n※フォントファイルコピー済み"
            messagebox.showinfo("PDF出力", msg)
            self._last_font_copied = False
        except Exception as e:
            messagebox.showerror("PDF出力エラー", str(e))
# 起動テスト用
if __name__ == "__main__":
    class DummySchedule:
        def __init__(self, date, start, end):
            self.date = date
            self.start_time = start
            self.end_time = end

    from datetime import date
    root = tk.Tk()
    root.withdraw()

    schedules = [
        DummySchedule(date(2025, 7, 20), "12:00", "20:00"),
        DummySchedule(date(2025, 7, 21), "13:00", "21:00")
    ]

    win = OrderWindow(root, schedules)
    win.mainloop()
