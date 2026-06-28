import csv, os, sys, json, shutil
import tkinter as tk
from tkinter import messagebox, ttk
from tkcalendar import DateEntry
from datetime import datetime, timedelta
import customtkinter as ctk
import config
from models import LiveSchedule, BandInfo

class TopWindow(ctk.CTkToplevel):
    def get_json_path(self, filename="schedule_data.json", mode="r"):
        if hasattr(sys, '_MEIPASS'):
            # PyInstallerでexe化した場合も、常にexeのある場所を参照
            base_dir = os.path.dirname(sys.executable)
        else:
            # スクリプト実行時はtop.pyのあるディレクトリ
            base_dir = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(base_dir, filename)

    def __init__(self, master=None):
        super().__init__(master)
        self.title("ライブ日程設定")
        self.geometry("840x650")  # 見出し追加に伴い高さを少し微調整
        self.minsize(800, 600)
        self.schedules = []
        self.band_data = self.load_band_infos_from_csv(os.path.join(os.path.dirname(sys.argv[0]), "bands.csv"))
        self.tabs = {}
        self.used_band_names = set()
        self.tab_control = None
        self.order_area = None
        self.top_widgets = []  # トップ画面ウィジェットを管理
        self.filter_options = {'オプション1': '', 'オプション2': '', 'オプション3': ''}  # フィルタ条件
        
        self.create_widgets()
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        self._last_font_copied = False
        
        # ウィンドウを最前面にフォーカスさせる
        self.after(200, lambda: self.focus())

    def _ensure_font_file(self, filename, friendly_name="フォント"):
        app_dir = os.path.dirname(os.path.abspath(__file__))
        target_path = os.path.join(app_dir, filename)
        if os.path.exists(target_path):
            self._last_font_copied = False
            return target_path

        windir = os.environ.get("WINDIR", "C:\\Windows")
        fonts_dir = os.path.join(windir, "Fonts")
        candidates = []
        try:
            for f in os.listdir(fonts_dir):
                if f.lower() == filename.lower() or filename.split('.')[0].lower() in f.lower():
                    candidates.append(os.path.join(fonts_dir, f))
        except Exception:
            candidates = []

        if not candidates:
            messagebox.showerror(f"{friendly_name} コピー失敗", f"{filename} が見つかりません。\nシステムフォントフォルダに {filename} が存在しないようです。", parent=self)
            return None

        for src in candidates:
            try:
                shutil.copy2(src, target_path)
                self._last_font_copied = True
                return target_path
            except Exception:
                continue

        messagebox.showerror(f"{friendly_name} コピー失敗", f"{filename} のコピーに失敗しました。手動で {filename} をアプリフォルダに置いてください。", parent=self)
        return None

    def on_close(self):
        if messagebox.askokcancel("確認", "変更内容を保存せずにウィンドウを閉じますか？", parent=self):
            self.destroy()

    def create_widgets(self):
        # 1. 画面大見出し (config.FONT_TITLE)
        title_label = ctk.CTkLabel(self, text="ライブ日程設定", font=config.FONT_TITLE)
        title_label.pack(pady=10, padx=15, anchor="w")
        self.top_widgets.append(title_label)

        # 2. 日程追加セクション (ctk.CTkFrame で再現)
        frame = ctk.CTkFrame(self)
        frame.pack(pady=10, padx=15, fill="x", anchor="w")
        self.top_widgets.append(frame)

        # セクションタイトル (config.FONT_SUBTITLE)
        sub_title1 = ctk.CTkLabel(frame, text="日程追加", font=config.FONT_SUBTITLE)
        sub_title1.grid(row=0, column=0, columnspan=3, padx=15, pady=5, sticky="w")

        ctk.CTkLabel(frame, text="日付：", font=config.FONT_LABEL_BUTTON, anchor="w").grid(row=1, column=0, padx=15, pady=5, sticky="w")
        # tkcalendarは標準のtkinterベースのためそのまま使用
        self.date_entry = DateEntry(frame, width=12, background='darkblue',
                                    foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd', font=config.FONT_LABEL_BUTTON)
        self.date_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")

        ctk.CTkLabel(frame, text="開始時刻：", font=config.FONT_LABEL_BUTTON, anchor="w").grid(row=2, column=0, padx=15, pady=5, sticky="w")
        self.start_time = ctk.CTkComboBox(frame, values=self.time_options(), width=120, font=config.FONT_LABEL_BUTTON)
        self.start_time.grid(row=2, column=1, padx=5, pady=5, sticky="w")

        ctk.CTkLabel(frame, text="終了時刻：", font=config.FONT_LABEL_BUTTON, anchor="w").grid(row=3, column=0, padx=15, pady=5, sticky="w")
        self.end_time = ctk.CTkComboBox(frame, values=self.time_options(), width=120, font=config.FONT_LABEL_BUTTON)
        self.end_time.grid(row=3, column=1, padx=5, pady=5, sticky="w")

        btn_add = ctk.CTkButton(frame, text="追加", command=self.add_schedule, font=config.FONT_LABEL_BUTTON, fg_color="#e0f7fa", text_color="black", hover_color="#b2dfdb", width=100)
        btn_add.grid(row=4, column=0, padx=15, pady=15, sticky="w")

        btn_load = ctk.CTkButton(frame, text="読み込み", command=self.load_schedules, fg_color="#b2dfdb", text_color="black", hover_color="#80cbc4", font=config.FONT_LABEL_BUTTON, width=100)
        btn_load.grid(row=4, column=1, padx=5, pady=15, sticky="w")

        btn_transfer = ctk.CTkButton(frame, text="他のPCへ引き継ぎ", command=self.transfer_data, font=config.FONT_LABEL_BUTTON, fg_color="#ffe082", text_color="black", hover_color="#ffd54f")
        btn_transfer.grid(row=4, column=2, padx=5, pady=15, sticky="w")

        # 3. 日程リスト表示エリア
        list_frame = ctk.CTkFrame(self)
        list_frame.pack(pady=10, padx=15, fill="both", expand=True, anchor="w")
        self.top_widgets.append(list_frame)

        sub_title2 = ctk.CTkLabel(list_frame, text="日程リスト", font=config.FONT_SUBTITLE)
        sub_title2.pack(padx=15, pady=5, anchor="w")

        list_inner_frame = ctk.CTkFrame(list_frame, fg_color="transparent")
        list_inner_frame.pack(fill="both", expand=True, padx=15, pady=5)

        # Listboxは複数行選択のロジック上、標準のtk.ListboxをCustomTkinterのフォントで装飾して使用
        self.listbox = tk.Listbox(list_inner_frame, font=config.FONT_LABEL_BUTTON, selectbackground="#ffe082", selectforeground="black", bd=0, highlightthickness=0)
        self.listbox.pack(side="left", fill="both", expand=True)

        scrollbar = ctk.CTkScrollbar(list_inner_frame, orientation="vertical", command=self.listbox.yview)
        scrollbar.pack(side="right", fill="y")
        self.listbox.config(yscrollcommand=scrollbar.set)

        # 4. ボタンエリア
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(pady=10, anchor="w", padx=15)
        self.top_widgets.append(btn_frame)

        btn_edit = ctk.CTkButton(btn_frame, text="編集", command=self.edit_schedule, fg_color="#ffe082", text_color="black", hover_color="#ffd54f", font=config.FONT_LABEL_BUTTON, width=80)
        btn_edit.pack(side="left", padx=8)

        btn_delete = ctk.CTkButton(btn_frame, text="削除", command=self.delete_schedule, fg_color="lightcoral", text_color="black", hover_color="#ff8a80", font=config.FONT_LABEL_BUTTON, width=80)
        btn_delete.pack(side="left", padx=8)

        btn_order = ctk.CTkButton(btn_frame, text="出演順設定エリアを表示", command=self.show_order_area, fg_color="lightblue", text_color="black", hover_color="#90caf9", font=config.FONT_LABEL_BUTTON, width=200)
        btn_order.pack(side="left", padx=8)

    def show_order_area(self):
        if not self.schedules:
            messagebox.showinfo("出演順設定", "日程を1つ以上追加してください", parent=self)
            return
        # トップ画面を非表示にする
        for widget in self.top_widgets:
            widget.pack_forget()
            
        if self.order_area:
            self.order_area.destroy()
        self.order_area = ctk.CTkFrame(self, fg_color="transparent")
        self.order_area.pack(fill="both", expand=True)
        self.create_order_area()

    def back_to_top(self):
        if self.order_area:
            self.order_area.destroy()
            self.order_area = None
            
        # トップ画面ウィジェットを元の設定で綺麗に再配置
        if len(self.top_widgets) >= 4:
            self.top_widgets[0].pack(pady=10, padx=15, anchor="w")  # 大見出し
            self.top_widgets[1].pack(pady=10, padx=15, fill="x", anchor="w")  # 日程追加
            self.top_widgets[2].pack(pady=10, padx=15, fill="both", expand=True, anchor="w")  # 日程リスト
            self.top_widgets[3].pack(pady=10, anchor="w", padx=15)  # ボタンエリア

    def create_order_area(self):
        menu_frame = ctk.CTkFrame(self.order_area, fg_color="transparent")
        menu_frame.pack(fill="x", padx=10, pady=3)
        
        btn_save = ctk.CTkButton(menu_frame, text="保存", command=self.save_schedule, font=config.FONT_LABEL_BUTTON, width=70)
        btn_save.pack(side="left", padx=10)
        
        btn_load = ctk.CTkButton(menu_frame, text="読み込み", command=self.load_schedule, font=config.FONT_LABEL_BUTTON, width=70)
        btn_load.pack(side="left", padx=10)
        
        btn_export_pdf = ctk.CTkButton(menu_frame, text="PDF出力", command=self.export_pdf, font=config.FONT_LABEL_BUTTON, width=80, fg_color="#ff417a", text_color="white", hover_color="#f50057")
        btn_export_pdf.pack(side="left", padx=10)
        
        btn_export_excel = ctk.CTkButton(menu_frame, text="Excel出力(おすすめ)", command=self.export_excel, font=config.FONT_LABEL_BUTTON, width=160, fg_color="#07ca6f", text_color="white", hover_color="#05964f")
        btn_export_excel.pack(side="left", padx=10)
        
        btn_option = ctk.CTkButton(menu_frame, text="絞り込み", command=self.show_option_dialog, font=config.FONT_LABEL_BUTTON, width=80, fg_color="#fff9c4", text_color="black", hover_color="#fff59d")
        btn_option.pack(side="left", padx=10)
        
        btn_back = ctk.CTkButton(menu_frame, text="日程設定に戻る", command=self.back_to_top, font=config.FONT_LABEL_BUTTON, width=120, fg_color="red", text_color="white", hover_color="#d32f2f")
        btn_back.pack(side="right", padx=10)

        # Tabviewへのアップグレード
        self.tab_control = ctk.CTkTabview(self.order_area)
        self.tab_control.pack(expand=True, fill="both", padx=10, pady=5)
        
        self.tabs = {}
        self.used_band_names = set()
        self.create_tabs()

    def show_option_dialog(self):
        win = ctk.CTkToplevel(self)
        win.title("絞り込みオプション")
        win.geometry("340x220")
        win.after(200, lambda: win.focus())
        
        labels = ["オプション1", "オプション2", "オプション3"]
        entries = {}
        for i, opt in enumerate(labels):
            ctk.CTkLabel(win, text=opt, font=config.FONT_LABEL_BUTTON).grid(row=i, column=0, padx=15, pady=8, sticky="e")
            ent = ctk.CTkEntry(win, width=180, font=config.FONT_LABEL_BUTTON)
            ent.insert(0, self.filter_options.get(opt, ''))
            ent.grid(row=i, column=1, padx=15, pady=8)
            entries[opt] = ent
            
        def on_ok():
            for opt in labels:
                self.filter_options[opt] = entries[opt].get().strip()
            win.destroy()
            for tab in self.tabs:
                self.refresh_combo(tab)
                
        def on_clear():
            for opt in labels:
                self.filter_options[opt] = ''
                entries[opt].delete(0, tk.END)
            win.destroy()
            for tab in self.tabs:
                self.refresh_combo(tab)
                
        btn_ok = ctk.CTkButton(win, text="OK", command=on_ok, font=config.FONT_LABEL_BUTTON, width=100, fg_color="#c8e6c9", text_color="black", hover_color="#a5d6a7")
        btn_ok.grid(row=4, column=0, pady=15, padx=15)
        
        btn_clear = ctk.CTkButton(win, text="全解除", command=on_clear, font=config.FONT_LABEL_BUTTON, width=100, fg_color="#ffcdd2", text_color="black", hover_color="#ef9a9a")
        btn_clear.grid(row=4, column=1, pady=15, padx=15)

    def load_band_infos_from_csv(self, csv_path):
        band_list = []
        try:
            with open(csv_path, newline='', encoding='utf-8') as csvfile:
                reader = csv.DictReader(csvfile)
                for row in reader:
                    name = row["バンド名"].strip()
                    minutes = int(row["演奏時間"].strip())
                    dates_str = row["出演日"].strip()
                    date_list = []
                    for part in dates_str.split(";"):
                        try:
                            date = datetime.strptime(part.strip(), "%Y-%m-%d").date()
                            date_list.append(date)
                        except ValueError:
                            pass
                    options = {k: row.get(k, '').strip() for k in ["オプション1", "オプション2", "オプション3"]}
                    other = row.get("その他", '').strip()
                    band = BandInfo(name, minutes, [d for d in date_list if isinstance(d, type(datetime.now().date()))])
                    band.options = options
                    band.other = other
                    band_list.append(band)
        except Exception:
            pass
        return band_list

    def create_tabs(self):
        for sched in self.schedules:
            label_text = sched['date'].strftime("%m/%d")
            self.tab_control.add(label_text)
            tab = self.tab_control.tab(label_text)
            
            top_frame = ctk.CTkFrame(tab, fg_color="transparent")
            top_frame.pack(fill="x", padx=10, pady=5)
            
            label = ctk.CTkLabel(top_frame, text="バンドを選択：", font=config.FONT_LABEL_BUTTON)
            label.pack(side="left")
            
            combo = ctk.CTkComboBox(top_frame, width=200, font=config.FONT_LABEL_BUTTON, state="readonly")
            combo.pack(side="left", padx=5)
            
            btn_add = ctk.CTkButton(top_frame, text="追加", command=lambda c=combo, t=label_text: self.add_band(c, t), font=config.FONT_LABEL_BUTTON, width=80, fg_color="#e0f7fa", text_color="black", hover_color="#b2dfdb")
            btn_add.pack(side="left", padx=5)
            
            btn_special = ctk.CTkButton(top_frame, text="特別枠追加", command=lambda t=label_text: self.add_special_frame(t), font=config.FONT_LABEL_BUTTON, width=100, fg_color="#ffe0b2", text_color="black", hover_color="#ffcc80")
            btn_special.pack(side="left", padx=5)
            
            label_info = ctk.CTkLabel(top_frame, text="バンド名をクリックして\nバンド情報を表示", font=config.FONT_LABEL_BUTTON, justify="left")
            label_info.pack(side="left", padx=5)
            
            # Canvasエリア一式を ctk.CTkScrollableFrame の1行に置き換え
            frame_container = ctk.CTkScrollableFrame(tab, fg_color="#f8f8f8")
            frame_container.pack(fill="both", expand=True, padx=5, pady=5)
            
            available_bands = [b for b in self.band_data if sched['date'] in b.available_dates]
            combo.configure(values=[b.name for b in available_bands])
            if available_bands:
                combo.set(available_bands[0].name)
            else:
                combo.set("")
                
            self.tabs[label_text] = {
                "combo": combo,
                "frame": frame_container,
                "bands": [],
                "band_objs": available_bands,
                "tab_label": label_text
            }

    def add_band(self, combo, tab):
        band_name = combo.get()
        if not band_name:
            return
        tab_info = self.tabs[tab]
        if any(b['type']=='band' and b['name']==band_name for b in tab_info["bands"]):
            messagebox.showwarning("重複追加", f"{band_name} はすでに他の日程で追加されています。", parent=self)
            return
        band = self.find_band_by_name(band_name, tab_info["band_objs"])
        if not band:
            messagebox.showerror("エラー", "バンド情報が見つかりません", parent=self)
            return
        tab_info["bands"].append({'type': 'band', 'name': band_name, 'minutes': band.performance_minutes})
        self.used_band_names.add(band_name)
        for t in self.tabs.keys():
            self.refresh_combo(t)
        self.update_band_frames(tab)

    def add_special_frame(self, tab):
        def on_ok():
            kind = var_kind.get()
            try:
                minutes = int(entry_min.get())
            except ValueError:
                messagebox.showerror("エラー", "分数は整数で入力してください", parent=self)
                return
            if kind == 'リハ':
                band_name = combo_band.get().strip()
                if not band_name:
                    messagebox.showerror("エラー", "バンド名を選択してください", parent=self)
                    return
                item = {'type': 'rehearsal', 'name': band_name, 'minutes': minutes}
            elif kind == '休憩':
                item = {'type': 'break', 'minutes': minutes}
            elif kind == '転換':
                item = {'type': 'change', 'minutes': minutes}
            else:
                return
            self.tabs[tab]["bands"].append(item)
            win.destroy()
            self.update_band_frames(tab)
            
        win = ctk.CTkToplevel(self)
        win.title("特別枠追加")
        win.geometry("320x180")
        win.after(200, lambda: win.focus())
        
        ctk.CTkLabel(win, text="種別", font=config.FONT_LABEL_BUTTON).grid(row=0, column=0, padx=10, pady=5)
        var_kind = tk.StringVar(value='休憩')
        kinds = ['休憩', '転換', 'リハ']
        combo_kind = ctk.CTkComboBox(win, values=kinds, variable=var_kind, state="readonly", width=120, font=config.FONT_LABEL_BUTTON)
        combo_kind.grid(row=0, column=1, padx=5, pady=5)
        
        ctk.CTkLabel(win, text="分数", font=config.FONT_LABEL_BUTTON).grid(row=1, column=0, padx=10, pady=5)
        entry_min = ctk.CTkEntry(win, width=120, justify='center', font=config.FONT_LABEL_BUTTON)
        entry_min.grid(row=1, column=1, padx=5, pady=5)
        
        ctk.CTkLabel(win, text="バンド名(リハのみ)", font=config.FONT_LABEL_BUTTON).grid(row=2, column=0, padx=10, pady=5)
        all_band_names = [b.name for b in self.tabs[tab]['band_objs']]
        combo_band = ctk.CTkComboBox(win, values=all_band_names, width=180, font=config.FONT_LABEL_BUTTON)
        combo_band.grid(row=2, column=1, padx=5, pady=5)
        
        btn = ctk.CTkButton(win, text="OK", command=on_ok, font=config.FONT_LABEL_BUTTON)
        btn.grid(row=3, column=0, columnspan=2, pady=15)
        
        def on_kind_change(choice):
            if choice == 'リハ':
                combo_band.configure(state='normal')
            else:
                combo_band.set("")
                combo_band.configure(state='disabled')
                
        combo_kind.configure(command=on_kind_change)
        on_kind_change(var_kind.get())

    def transfer_data(self):
        result = messagebox.askyesno("確認", "引き継ぎ先のコンピュータに出席管理システムがインストールされていますか？", parent=self)
        if result:
            self.save_schedule()
            desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
            target_dir = os.path.join(desktop, '【ロック部出席管理】引き継ぎデータ')
            os.makedirs(target_dir, exist_ok=True)
            files = [
                os.path.join(os.path.dirname(sys.argv[0]), 'attend_data.xlsx'),
                os.path.join(os.path.dirname(sys.argv[0]), 'bands.csv'),
                os.path.join(os.path.dirname(sys.argv[0]), 'schedule_data.json'),
            ]
            copied = []
            for f in files:
                if os.path.exists(f):
                    shutil.copy2(f, target_dir)
                    copied.append(os.path.basename(f))
            msg = f"編集中の日程・出演順データを保存し、引き継ぎデータをデスクトップの『【ロック部出席管理】引き継ぎデータ』フォルダに保存しました。\n\n" \
                  f"引き継ぎ先のコンピュータにこのフォルダごと移動すると引き継ぎが完了します。\n\n" \
                  f"保存ファイル: {', '.join(copied)}"
            messagebox.showinfo("引き継ぎ完了", msg, parent=self)
        else:
            messagebox.showinfo("案内", "引き継ぎ先のコンピュータにソフトウェアをインストールしてから引き継ぎを行ってください。", parent=self)

    def update_band_frames(self, tab):
        tab_info = self.tabs[tab]
        frame = tab_info["frame"]
        for widget in frame.winfo_children():
            widget.destroy()
        sched = None
        tab_label = tab_info.get("tab_label")
        for s in self.schedules:
            if s['date'].strftime("%m/%d") == tab_label:
                sched = s
                break
        if not sched:
            return
        end_time_limit = datetime.strptime(sched['end'], "%H:%M")
        current_time = datetime.strptime(sched['start'], "%H:%M")
        
        for idx, item in enumerate(tab_info["bands"]):
            if item['type'] == 'band':
                name = item['name']
                minutes = item['minutes']
                label_text = f"{name}（{minutes}分）"
                band_info = self.find_band_by_name(name, self.band_data)
                def show_band_info(event, band_info=band_info):
                    if band_info:
                        info = f"バンド名: {band_info.name}\n演奏時間: {band_info.performance_minutes}分\n出演日: {', '.join([d.strftime('%Y-%m-%d') for d in band_info.available_dates])}"
                        if hasattr(band_info, 'options'):
                            for k, v in band_info.options.items():
                                info += f"\n{k}: {v}"
                        if hasattr(band_info, 'other') and band_info.other:
                            info += f"\nその他: {band_info.other}"
                        messagebox.showinfo("バンド情報", info, parent=self)
            elif item['type'] == 'break':
                name = None
                minutes = item['minutes']
                label_text = f"休憩（{minutes}分）"
            elif item['type'] == 'change':
                name = None
                minutes = item['minutes']
                label_text = f"転換（{minutes}分）"
            elif item['type'] == 'rehearsal':
                name = item['name']
                minutes = item['minutes']
                label_text = f"リハ（{name}）（{minutes}分）"
            else:
                continue
                
            start_str = current_time.strftime("%H:%M")
            end_time = current_time + timedelta(minutes=minutes)
            end_str = end_time.strftime("%H:%M")
            
            bg_color = None
            if item['type'] == 'band' and band_info is not None:
                other_val = getattr(band_info, 'other', None)
                if other_val is not None and str(other_val).lower() != 'nan' and str(other_val).strip() != '':
                    bg_color = '#b2ebf2'  # 水色
            if bg_color is None and end_time > end_time_limit:
                bg_color = '#fff59d'  # 黄色
                
            # ctk.CTkFrame でのコンテナ
            if bg_color:
                wrapper = ctk.CTkFrame(frame, corner_radius=4, border_width=1, border_color="gray", fg_color=bg_color, height=40)
            else:
                wrapper = ctk.CTkFrame(frame, corner_radius=4, border_width=1, border_color="gray", height=40)
                
            wrapper.pack(fill="x", expand=True, pady=2, padx=5)
            wrapper.pack_propagate(False)
            
            label = ctk.CTkLabel(wrapper, text=f"{start_str}～{end_str} {label_text}", anchor="w", font=config.FONT_LABEL_BUTTON, text_color="black" if bg_color else None)
            label.pack(side="left", fill="x", expand=True, padx=10)
            
            if item['type'] == 'band':
                label.bind('<Button-1>', show_band_info)
                
            btn_del = ctk.CTkButton(wrapper, text="×", width=30, height=25, fg_color="transparent", text_color="red", hover_color="lightcoral", font=config.FONT_LABEL_BUTTON)
            btn_del.pack(side="right", padx=5)
            
            btn_down = ctk.CTkButton(wrapper, text="↓", width=25, height=25, font=config.FONT_LABEL_BUTTON)
            btn_down.pack(side="right", padx=2)
            
            btn_up = ctk.CTkButton(wrapper, text="↑", width=25, height=25, font=config.FONT_LABEL_BUTTON)
            btn_up.pack(side="right", padx=2)
            
            def remove_item(idx=idx):
                removed = tab_info["bands"][idx]
                del tab_info["bands"][idx]
                if removed.get("type") == "band":
                    self.used_band_names.discard(removed["name"])
                    for t in self.tabs:
                        self.refresh_combo(t)
                self.update_band_frames(tab)
                
            btn_up.configure(command=lambda idx=idx: self.move_band(tab, idx, -1))
            btn_down.configure(command=lambda idx=idx: self.move_band(tab, idx, 1))
            btn_del.configure(command=remove_item)
            
            if end_time > end_time_limit:
                warn = ctk.CTkLabel(wrapper, text="時刻超過", fg_color='#fff59d', text_color="black", font=config.FONT_LABEL_BUTTON)
                warn.pack(side="right", padx=10)
            current_time = end_time

    def move_band(self, tab, idx, direction):
        tab_info = self.tabs[tab]
        new_idx = idx + direction
        if 0 <= new_idx < len(tab_info["bands"]):
            tab_info["bands"][idx], tab_info["bands"][new_idx] = tab_info["bands"][new_idx], tab_info["bands"][idx]
            self.update_band_frames(tab)

    def refresh_combo(self, tab):
        tab_info = self.tabs[tab]
        used_names = self.used_band_names
        def match_option(band):
            for opt, val in self.filter_options.items():
                v = val.strip()
                if v == '':
                    continue
                band_opt = band.options.get(opt, '').strip() if hasattr(band, 'options') else ''
                if band_opt != v:
                    return False
            return True
        filtered = [b.name for b in tab_info["band_objs"] if b.name not in used_names and match_option(b)]
        if all(v.strip() == '' for v in self.filter_options.values()):
            filtered = [b.name for b in tab_info["band_objs"] if b.name not in used_names]
            
        tab_info["combo"].configure(values=filtered)
        if filtered:
            tab_info["combo"].set(filtered[0])
        else:
            tab_info["combo"].set("")

    def find_band_by_name(self, name, band_list):
        for b in band_list:
            if b.name == name:
                return b
        return None

    def save_schedule(self):
        data = {
            "schedules": [
                {
                    "date": s['date'].strftime("%Y-%m-%d"),
                    "start_time": s['start'],
                    "end_time": s['end']
                } for s in self.schedules
            ],
            "bands": {
                tabinfo.get("tab_label", str(i)): tabinfo["bands"]
                for i, (tab, tabinfo) in enumerate(self.tabs.items())
            }
        }
        try:
            json_path = self.get_json_path()
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
                messagebox.showinfo("保存", "スケジュールを保存しました。", parent=self)
        except Exception as e:
                messagebox.showerror("保存エラー", str(e), parent=self)

    def load_schedule(self):
        try:
            json_path = self.get_json_path()
            with open(json_path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            messagebox.showerror("読み込みエラー", str(e), parent=self)
            return
        date2tab = {tabinfo.get("tab_label"): tab for tab, tabinfo in self.tabs.items()}
        bands_data = data.get("bands", {})
        self.used_band_names.clear()
        for tab_label, bands in bands_data.items():
            tab = date2tab.get(tab_label)
            if tab:
                self.tabs[tab]["bands"] = bands
        for tab in self.tabs:
            for b in self.tabs[tab]["bands"]:
                if b.get("type") == "band":
                    self.used_band_names.add(b["name"])
        for tab in self.tabs:
            self.refresh_combo(tab)
            self.update_band_frames(tab)
        messagebox.showinfo("読み込み", "スケジュールを読み込みました。", parent=self)

    def export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
        except ImportError:
            messagebox.showerror("Excel出力エラー", "openpyxlライブラリが必要です。\npip install openpyxl でインストールしてください。", parent=self)
            return
        font_file = os.path.join(os.path.dirname(__file__), "meiryo.ttc")
        if not os.path.exists(font_file):
            if not self._ensure_font_file("meiryo.ttc", "Meiryo"):
                return

        wb = openpyxl.Workbook()
        for tab, tabinfo in self.tabs.items():
            tab_title = tabinfo.get("tab_label")
            safe_title = tab_title.replace("/", "-")
            ws = wb.create_sheet(title=safe_title)
            ws.append(["開始時刻", "～", "終了時刻", "枠の名前"])
            header_font = Font(name="Meiryo", size=12, bold=True)
            for col in range(1, 5):
                ws.cell(row=ws.max_row, column=col).font = header_font
            for idx, item in enumerate(tabinfo["bands"]):
                sched = None
                for s in self.schedules:
                    if s['date'].strftime("%m/%d") == tabinfo.get("tab_label"):
                        sched = s
                        break
                if not sched:
                    continue
                current_time = datetime.strptime(sched['start'], "%H:%M")
                for i in range(idx):
                    current_time += timedelta(minutes=tabinfo["bands"][i]["minutes"])
                start_str = current_time.strftime("%H:%M")
                end_time = current_time + timedelta(minutes=item["minutes"])
                end_str = end_time.strftime("%H:%M")
                name = item.get("name") if item.get("type") != "break" and item.get("type") != "change" else None
                if item["type"] == "band":
                    row = [start_str, "～", end_str, name]
                    ws.append(row)
                    for col in range(1, 5):
                        ws.cell(row=ws.max_row, column=col).fill = PatternFill(fill_type=None)
                        ws.cell(row=ws.max_row, column=col).font = Font(name="Meiryo", size=11, color="000000")
                else:
                    if item["type"] == "break":
                        disp_name = f"休憩（{item['minutes']}分）"
                    elif item["type"] == "change":
                        disp_name = f"転換（{item['minutes']}分）"
                    elif item["type"] == "rehearsal":
                        disp_name = f"リハ（{item['name']}）（{item['minutes']}分）"
                    else:
                        disp_name = ""
                    row = [start_str, "～", end_str, disp_name]
                    ws.append(row)
                    for col in range(1, 5):
                        ws.cell(row=ws.max_row, column=col).fill = PatternFill(start_color="888888", end_color="888888", fill_type="solid")
                        ws.cell(row=ws.max_row, column=col).font = Font(name="Meiryo", size=11, color="FFFFFF")
                for col in range(1, 5):
                    ws.cell(row=ws.max_row, column=col).alignment = Alignment(horizontal="left", vertical="center")
            ws.column_dimensions["A"].width = 10
            ws.column_dimensions["B"].width = 4
            ws.column_dimensions["C"].width = 10
            ws.column_dimensions["D"].width = 24
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        try:
            wb.save("タイムテーブル.xlsx")
            msg = "タイムテーブル.xlsx を出力しました。"
            if getattr(self, '_last_font_copied', False):
                msg += "\n※フォントファイルコピー済み"
            messagebox.showinfo("Excel出力", msg, parent=self)
            self._last_font_copied = False
        except Exception as e:
            messagebox.showerror("Excel出力エラー", str(e), parent=self)

    def export_pdf(self):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
            from reportlab.lib.colors import black, white, HexColor
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            font_path = os.path.join(os.path.dirname(__file__), "meiryo.ttc")
            if not os.path.exists(font_path):
                if not self._ensure_font_file("meiryo.ttc", "Meiryo"):
                    return
            pdfmetrics.registerFont(TTFont("meiryo", font_path))
        except ImportError:
            messagebox.showerror("PDF出力エラー", "reportlabライブラリが必要です。\npip install reportlab でインストールしてください。", parent=self)
            return
        try:
            c = canvas.Canvas("タイムテーブル.pdf", pagesize=A4)
            width, height = A4
            y_start = height - 50
            for tab, tabinfo in self.tabs.items():
                c.setFont("meiryo", 16)
                c.drawString(50, y_start, f"日程: {tabinfo.get('tab_label')}")
                y = y_start - 30
                c.setFont("meiryo", 12)
                for idx, item in enumerate(tabinfo["bands"]):
                    sched = None
                    for s in self.schedules:
                        if s['date'].strftime("%m/%d") == tabinfo.get('tab_label'):
                            sched = s
                            break
                    if not sched:
                        continue
                    current_time = datetime.strptime(sched['start'], "%H:%M")
                    for i in range(idx):
                        current_time += timedelta(minutes=tabinfo["bands"][i]["minutes"])
                    start_str = current_time.strftime("%H:%M")
                    end_time = current_time + timedelta(minutes=item["minutes"])
                    end_str = end_time.strftime("%H:%M")
                    if item["type"] == "band":
                        c.setFillColor(white)
                        c.setStrokeColor(black)
                        c.rect(45, y-2, 340, 22, fill=1, stroke=0)
                        c.setFillColor(black)
                        c.drawString(50, y, start_str)
                        c.drawString(120, y, "～")
                        c.drawString(160, y, end_str)
                        c.drawString(250, y, item["name"])
                    else:
                        if item["type"] == "break":
                            disp_name = f"休憩（{item['minutes']}分）"
                        elif item["type"] == "change":
                            disp_name = f"転換（{item['minutes']}分）"
                        elif item["type"] == "rehearsal":
                            disp_name = f"リハ（{item['name']}）（{item['minutes']}分）"
                        else:
                            disp_name = ""
                        c.setFillColor(HexColor("#888888"))
                        c.setStrokeColor(black)
                        c.rect(45, y-2, 340, 22, fill=1, stroke=0)
                        c.setFillColor(white)
                        c.drawString(50, y, start_str)
                        c.drawString(120, y, "～")
                        c.drawString(160, y, end_str)
                        c.drawString(250, y, disp_name)
                    y -= 24
                y_start = y - 40
                c.showPage()
            c.save()
            msg = "タイムテーブル.pdf を出力しました。"
            if getattr(self, '_last_font_copied', False):
                msg += "\n※フォントファイルコピー済み"
            messagebox.showinfo("PDF出力", msg, parent=self)
            self._last_font_copied = False
        except Exception as e:
            messagebox.showerror("PDF出力エラー", str(e), parent=self)

    def time_options(self):
        return [f"{h:02d}:{m:02d}" for h in range(8, 21) for m in (0, 30)]

    def add_schedule(self):
        date = self.date_entry.get_date()
        start = self.start_time.get()
        end = self.end_time.get()

        if not self.is_valid_time_format(start) or not self.is_valid_time_format(end):
            messagebox.showerror("エラー", "開始時刻・終了時刻はHH:MM形式で入力してください", parent=self)
            return

        start_dt = datetime.strptime(start, "%H:%M")
        end_dt = datetime.strptime(end, "%H:%M")
        if end_dt <= start_dt:
            messagebox.showerror("エラー", "終了時刻は開始時刻より後にしてください", parent=self)
            return

        for s in self.schedules:
            if s['date'] == date:
                messagebox.showwarning("重複", "同じ日付が既に追加されています", parent=self)
                return

        self.schedules.append({
            "date": date,
            "start": start,
            "end": end
        })

        self.update_listbox()

    def delete_schedule(self):
        selected = self.listbox.curselection()
        if not selected:
            messagebox.showinfo("削除", "削除する日程を選択してください", parent=self)
            return

        index = selected[0]
        del self.schedules[index]
        self.update_listbox()

    def edit_schedule(self):
        selected = self.listbox.curselection()
        if not selected:
            messagebox.showinfo("編集", "編集する日程を選択してください", parent=self)
            return
        idx = selected[0]
        s = self.schedules[idx]
        
        edit_win = ctk.CTkToplevel(self)
        edit_win.title("日程編集")
        edit_win.geometry("320x220")
        edit_win.after(200, lambda: edit_win.focus())
        
        ctk.CTkLabel(edit_win, text="日付：", font=config.FONT_LABEL_BUTTON).grid(row=0, column=0, padx=10, pady=10)
        date_entry = DateEntry(edit_win, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd', font=config.FONT_LABEL_BUTTON)
        date_entry.set_date(s['date'])
        date_entry.grid(row=0, column=1, padx=10, pady=10)
        
        ctk.CTkLabel(edit_win, text="開始時刻：", font=config.FONT_LABEL_BUTTON).grid(row=1, column=0, padx=10, pady=10)
        start_combo = ctk.CTkComboBox(edit_win, values=self.time_options(), width=120, font=config.FONT_LABEL_BUTTON)
        start_combo.set(s['start'])
        start_combo.grid(row=1, column=1, padx=10, pady=10)
        
        ctk.CTkLabel(edit_win, text="終了時刻：", font=config.FONT_LABEL_BUTTON).grid(row=2, column=0, padx=10, pady=10)
        end_combo = ctk.CTkComboBox(edit_win, values=self.time_options(), width=120, font=config.FONT_LABEL_BUTTON)
        end_combo.set(s['end'])
        end_combo.grid(row=2, column=1, padx=10, pady=10)
        
        def save_edit():
            new_date = date_entry.get_date()
            new_start = start_combo.get()
            new_end = end_combo.get()
            if not self.is_valid_time_format(new_start) or not self.is_valid_time_format(new_end):
                messagebox.showerror("エラー", "開始時刻・終了時刻はHH:MM形式で入力してください", parent=self)
                return
            start_dt = datetime.strptime(new_start, "%H:%M")
            end_dt = datetime.strptime(new_end, "%H:%M")
            if end_dt <= start_dt:
                messagebox.showerror("エラー", "終了時刻は開始時刻より後にしてください", parent=self)
                return
            for i, ss in enumerate(self.schedules):
                if i != idx and ss['date'] == new_date:
                    messagebox.showwarning("重複", "同じ日付が既に追加されています", parent=self)
                    return
            self.schedules[idx] = {"date": new_date, "start": new_start, "end": new_end}
            self.update_listbox()
            edit_win.destroy()
            
        btn_save = ctk.CTkButton(edit_win, text="保存", command=save_edit, fg_color="#e0f7fa", text_color="black", hover_color="#b2dfdb", font=config.FONT_LABEL_BUTTON)
        btn_save.grid(row=3, column=0, columnspan=2, pady=15)

    def update_listbox(self):
        self.listbox.delete(0, tk.END)
        for s in self.schedules:
            date_str = s['date'].strftime('%Y-%m-%d')
            self.listbox.insert(tk.END, f"{date_str} / {s['start']}～{s['end']}")

    def is_valid_time_format(self, time_str):
        try:
            datetime.strptime(time_str, "%H:%M")
            return True
        except ValueError:
            return False

    def load_schedules(self):
        try:
            json_path = self.get_json_path()
            with open(json_path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            messagebox.showerror("読み込みエラー", str(e), parent=self)
            return
        self.schedules.clear()
        for s in data.get("schedules", []):
            try:
                d = datetime.strptime(s["date"], "%Y-%m-%d").date()
                start = s["start_time"] if "start_time" in s else s["start"]
                end = s["end_time"] if "end_time" in s else s["end"]
                self.schedules.append({"date": d, "start": start, "end": end})
            except Exception:
                continue
        self.update_listbox()
        messagebox.showinfo("読み込み", "日程を読み込みました。", parent=self)


# 起動確認用のテストブロック
if __name__ == "__main__":
    ctk.set_appearance_mode("System")
    ctk.set_default_color_theme("blue")
    
    # 単体動作テスト用にメインウィンドウを非表示で生成
    root = ctk.CTk()
    root.withdraw()
    
    app = TopWindow(root)
    root.mainloop()