# 2026年6月19日更新
# 出席率計算を行うクラス・関数をここに分割して記述

import pandas as pd, os
import openpyxl, datetime
from tkinter import messagebox

def is_date_in_range(date_str, start_date, end_date):
    """日付が指定期間内かどうか判定（簡易版）"""
    try:
        # M/D形式の日付を月・日に分割
        def parse_date(date_string):
            parts = date_string.strip().split('/')
            if len(parts) == 2:
                return int(parts[0]), int(parts[1])
            return None, None

        target_month, target_day = parse_date(date_str)
        start_month, start_day = parse_date(start_date)
        end_month, end_day = parse_date(end_date)

        if None in [target_month, target_day, start_month, start_day, end_month, end_day]:
            return False

        target_score = target_month * 100 + target_day
        start_score = start_month * 100 + start_day
        end_score = end_month * 100 + end_day

        # 年跨ぎ対応: 開始日 > 終了日なら年を跨ぐとみなす
        if start_score <= end_score:
            # 通常: 開始日 <= target <= 終了日
            return start_score <= target_score <= end_score
        else:
            # 年跨ぎ: targetが開始日以降 or 終了日以前
            return (target_score >= start_score) or (target_score <= end_score)
    except:
        return False

def save_attendance_rate_to_excel(results, file_path):
    """出席率記録シートに期間ごとに記録（出欠情報と同様の形式・書式コピーあり）"""
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet_name = '出席率記録'
        if sheet_name not in wb.sheetnames:
            messagebox.showerror('エラー', f'出席率を記録するシートが存在しません。\n適切なファイルを指定しているか確認してください。')
        else:
            ws = wb[sheet_name]

        # 期間列（5列目以降）で、今回の期間が既にあればその列、なければ空白列または末尾
        period = results[0]['期間'] if results else ''
        target_col = None
        max_col = ws.max_column
        # まず既存の期間列を探す
        for col in range(7, max_col + 1):
            val = ws.cell(row=2, column=col).value
            if str(val) == period:
                target_col = col
                break
        if target_col is None:
            # 空白列を探す
            for col in range(7, max_col + 1):
                val = ws.cell(row=2, column=col).value
                if val is None or str(val).strip() == '':
                    target_col = col
                    break
        if target_col is None:
            target_col = max_col + 1

        # ヘッダが空なら期間を設定
        period_cell = ws.cell(row=2, column=target_col)
        if period_cell.value is None or str(period_cell.value).strip() == '':
            period_cell.value = period

        # 各行に出席率を書き込む
        for i, result in enumerate(results):
            row_num = 3 + i
            cell = ws.cell(row=row_num, column=target_col)
            cell.value = result['出席率']
        wb.save(file_path)
    except Exception as e:
        raise Exception(f'Excelへの保存に失敗しました: {str(e)}')

def calculate_attendance_rate(start_date, end_date, file_path, sheet_name):
    """指定期間の出席率を計算し、出席率記録シートに保存"""
    try:
        # 入力検証
        if not start_date or not end_date:
            messagebox.showerror('エラー', '開始日と終了日を入力してください')
            return
        
        # 出欠状況シートを読み込み
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=1, index_col=None)

        # 日付列を特定（G列以降）
        date_columns = []
        for col in df.columns[6:]:  # G列以降
            if pd.notna(col) and str(col).strip():
                date_columns.append(col)
        
        # 指定期間内の日付列を抽出
        target_columns = []
        for col in date_columns:
            # 簡易的な期間判定（月/日形式）
            if is_date_in_range(str(col), start_date, end_date):
                target_columns.append(col)
        
        if not target_columns:
            messagebox.showwarning('警告', '指定期間内にデータが見つかりません')
            return
        
        # 各学生の出席率を計算
        results = []
        for idx, row in df.iterrows():
            if pd.isna(row.get('氏名')) or not str(row.get('氏名')).strip():
                break  # 氏名が空白の行で終了
            
            # 出欠データを集計
            attendance_count = 0  # 出席をとった回数
            absent_with_contact = 0  # 連絡あり欠席（△）
            absent_without_contact = 0  # 無断欠席（×）
            
            # 連絡あり欠席：-1、無断欠席：-2として計算、オンライン・忌引きは出席扱い
            for col in target_columns:
                value = str(row.get(col, '')).strip()
                if value in ['出席', '連絡あり', '無断欠席', 'オ', '忌引', '']:
                    attendance_count += 1
                    if value == '連絡あり':
                        absent_with_contact += 1
                    elif value == '無断欠席':
                        absent_without_contact += 1
            
            # 出席率計算（マイナス値もそのまま記録、小数第4位まで）
            if attendance_count > 0:
                rate = 100 * (attendance_count - absent_with_contact - absent_without_contact * 2) / attendance_count
            else:
                rate = 0

            results.append({
                '学年': row.get('学年', ''),
                '学部': row.get('学部', ''),
                '学籍番号': row.get('学籍番号', ''),
                '氏名': row.get('氏名', ''),
                '出席率': round(rate, 4),
                '期間': f"{start_date}～{end_date}"
            })
        
        # 出席率記録シートに保存
        save_attendance_rate_to_excel(results, file_path)
        
        messagebox.showinfo('完了', f'出席率を計算しました。\n対象期間: {start_date}～{end_date}\n対象人数: {len(results)}人')
        
    except Exception as e:
        messagebox.showerror('エラー', f'出席率の計算中にエラーが発生しました:\n{str(e)}')

def calculate_rate_and_export(start_date, end_date, file_path, sheet_name):
    """指定期間の出席率を計算し、テキストファイルに出力"""
    try:
        # 入力検証
        if not start_date or not end_date:
            messagebox.showerror('エラー', '開始日と終了日を入力してください')
            return
        
        # 出欠状況シートを読み込み
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=1, index_col=None)

        # 日付列を特定（G列以降）
        date_columns = []
        for col in df.columns[6:]:  # G列以降
            if pd.notna(col) and str(col).strip():
                date_columns.append(col)
        
        # 指定期間内の日付列を抽出
        target_columns = []
        for col in date_columns:
            # 簡易的な期間判定（月/日形式）
            if is_date_in_range(str(col), start_date, end_date):
                target_columns.append(col)
        
        if not target_columns:
            messagebox.showwarning('警告', '指定期間内にデータが見つかりません')
            return
        
        # 各学生の出席率を計算
        results = []
        for idx, row in df.iterrows():
            if pd.isna(row.get('氏名')) or not str(row.get('氏名')).strip():
                break  # 氏名が空白の行で終了
            
            # 出欠データを集計
            attendance_count = 0  # 出席をとった回数
            absent_with_contact = 0  # 連絡あり欠席（△）
            absent_without_contact = 0  # 無断欠席（×）
            
            # 連絡あり欠席：-1、無断欠席：-2として計算、オンライン・忌引きは出席扱い
            for col in target_columns:
                value = str(row.get(col, '')).strip()
                if value in ['出席', '連絡あり', '無断欠席', 'オ', '忌引', '']:
                    attendance_count += 1
                    if value == '連絡あり':
                        absent_with_contact += 1
                    elif value == '無断欠席':
                        absent_without_contact += 1
            
            # 出席率計算（マイナス値もそのまま記録、小数第4位まで）
            if attendance_count > 0:
                rate = 100 * (attendance_count - absent_with_contact - absent_without_contact * 2) / attendance_count
            else:
                rate = 0

            results.append({
                '学年': row.get('学年', ''),
                '学部': row.get('学部', ''),
                '学籍番号': row.get('学籍番号', ''),
                '氏名': row.get('氏名', ''),
                '出席率': round(rate, 4),
                '期間': f"{start_date}～{end_date}"
            })
                
        # テキストファイルに出力(デスクトップに保存)
        desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
        with open(os.path.join(desktop_path, 'attendance_rates.txt'), 'w', encoding='utf-8') as f:
            f.write(f"出席率 - 対象期間: {start_date}～{end_date}\n")
            for result in results:
                f.write(f"{result['氏名']} ({result['学籍番号']}): {result['出席率']}%\n")
            f.write(f"\n出力: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        
        messagebox.showinfo('完了', f'出席率を計算しました。\n対象期間: {start_date}～{end_date}\n対象人数: {len(results)}人')
        
    except Exception as e:
        messagebox.showerror('エラー', f'出席率の計算中にエラーが発生しました:\n{str(e)}')