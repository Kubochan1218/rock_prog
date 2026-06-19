# 2026年6月19日更新

import datetime, openpyxl, re, os, sys
import tkinter as tk
from tkinter import messagebox, simpledialog, filedialog, ttk
import tkinter.font as tkfont
import pandas as pd
import band_selection as bs
import attendance_calculation as ac
from top import TopWindow

FILE_PATH = 'attend_data.xlsx'
SHEET_NAME = '出欠状況'
FONT_NAME = 'Yu Gothic UI'


class AttendanceApp:
    def show_top(self):
        self.clear()
        tk.Label(self.master, text='ロック部 出席管理', font=(FONT_NAME, 16, 'bold')).pack(pady=10)
        # 操作支援が有効ならば簡易ヘルプを表示
        try:
            if self.settings.get('operation_support', False):
                tk.Label(self.master, text='操作支援: 設定が有効です。ボタンにマウスを合わせると説明が表示されます。\n※「設定」で操作支援のON/OFFを切り替えられます。', font=(FONT_NAME, 10), fg='#333333').pack(pady=4)
        except Exception:
            pass
        # ヘルプモード用レイアウト（操作支援が有効かつ help_mode が ON の場合）
        try:
            if self.settings.get('operation_support', False) and self.settings.get('help_mode', False):
                hl_frame = tk.Frame(self.master)
                hl_frame.pack(pady=4, fill='both', expand=True)
                #tk.Label(hl_frame, text='上から順に操作してください。', font=(FONT_NAME, 14, 'bold')).pack(pady=6)
                items = [
                    ('① 出欠をとる', '開く', self.show_attendance_date_select, '#bfff80'),
                    ('② 出席率の計算', '開く', self.show_attendance_check, '#80d4ff'),
                    ('③ バンド登録', '開く', self.register_band, '#ffff00'),
                    ('④ バンド選出', '開く', self.show_select_band, '#ffd480'),
                    ('⑤ タイムテーブル作成', '開く', self.make_timetable, '#d080ff'),
                ]
                # 進捗インデックス（何個目まで実行済みか）
                prog = getattr(self, 'help_progress_index', 0)
                def make_help_cmd(idx, fn):
                    def inner():
                        try:
                            # 実行後に進捗を更新
                            try:
                                fn()
                            except Exception:
                                pass
                            self.help_progress_index = max(getattr(self, 'help_progress_index', 0), idx + 1)
                        except Exception:
                            pass
                    return inner

                for i, (title, btn_text, cmd, default_color) in enumerate(items):
                    # 状態判定: 実行済み(<prog)=緑, 次に実行(==prog)=青, 表示保留(>prog)=灰
                    if i < prog:
                        color = "#00cf4f"  # 実行済み（緑）
                    elif i == prog:
                        color = "#174dff"  # 次に実行（青）
                    else:
                        color = '#cccccc'  # 表示保留（灰）
                        default_color = '#cccccc'

                    row = tk.Frame(hl_frame)
                    row.pack(fill='x', padx=12, pady=12, ipady=6)
                    lbl = tk.Label(row, text=title, font=(FONT_NAME, 12), fg=color)
                    lbl.pack(side='left')

                    btn = tk.Button(row, text=btn_text, width=12, height=1, bg=default_color, font=(FONT_NAME, 12, 'bold'), command=make_help_cmd(i, cmd))
                    btn.place(relx=0.0, rely=0.5, anchor='w', x=180)
                    try:
                        self.add_tooltip(btn, f'{title} の画面を開きます')
                    except Exception:
                        pass
                # 設定ボタン（左下に固定配置）
                btn_settings = tk.Button(
                self.master,
                text='設定',
                width=10,
                height=1,
                command=self.show_settings,
                bg='#cccccc',  # グレー
                font=(FONT_NAME, 16, 'bold')
                )
                btn_settings.place(relx=0.0, rely=1.0, anchor='sw', x=20, y=-20) # 左下に固定配置
                # 設定ボタンにもツールチップ
                self.add_tooltip(btn_settings, 'アプリの各種設定を開きます')

                # ヘルプモード終了ボタン
                btn_exit = tk.Button(self.master, text='ヘルプモードを終了', width=18, height=1, bg='#cccccc', font=(FONT_NAME, 12, 'bold'), command=self.toggle_help_mode)
                btn_exit.place(relx=1.0, rely=1.0, anchor='se', x=-20, y=-20)
                try:
                    self.add_tooltip(btn_exit, 'ヘルプモードを終了します')
                except Exception:
                    pass
                return
        except Exception:
            pass
        # 上段ボタンフレーム
        btn_frame_top = tk.Frame(self.master)
        btn_frame_top.pack(pady=10)
        btn_attendance = tk.Button(
            btn_frame_top,
            text='出欠をとる',
            width=14,
            height=2,
            command=self.show_attendance_date_select,
            bg='#bfff80',  # 黄緑色
            font=(FONT_NAME, 16, 'bold')
        )
        btn_attendance.pack(side='left', padx=10, ipadx=10, ipady=8)
        self.add_tooltip(btn_attendance, '出欠を記録する画面を開きます')

        btn_check = tk.Button(
            btn_frame_top,
            text='出欠状況の確認\n(出席率計算)',
            width=14,
            height=2,
            command=self.show_attendance_check,
            bg='#80d4ff',  # 水色
            font=(FONT_NAME, 16, 'bold')
        )
        btn_check.pack(side='left', padx=10, ipadx=10, ipady=8)
        self.add_tooltip(btn_check, '出席率の計算と出欠状況の確認を行います')

        btn_frame_register = tk.Frame(btn_frame_top)
        btn_frame_register.pack(pady=10)

        btn_register = tk.Button(
            btn_frame_register,
            text='バンドを登録',
            width=14,
            height=1,
            command=self.register_band,
            bg='#ffff00',  # 黄色
            font=(FONT_NAME, 16, 'bold')
        )
        btn_register.pack(padx=10, pady=1, ipadx=10, ipady=2)
        self.add_tooltip(btn_register, '新しいバンドを登録します')
        
        # 登録済みバンドをクリアするボタン
        btn_clear_bands = tk.Button(
            btn_frame_register,
            text='登録済みバンドクリア',
            width=14,
            height=1,
            command=self.clear_registered_bands,
            bg='#ff6666',
            font=(FONT_NAME, 12)
        )
        btn_clear_bands.pack(padx=10, pady=1, ipadx=31, ipady=1)
        try:
            self.add_tooltip(btn_clear_bands, '登録済みのバンドをすべてクリアします')
        except Exception:
            pass

        # 下段ボタンフレーム
        btn_frame_bottom = tk.Frame(self.master)
        btn_frame_bottom.pack(pady=5)
        btn_select = tk.Button(
            btn_frame_bottom,
            text='出演バンド選出',
            width=14,
            height=2,
            command=self.show_select_band,
            bg='#ffd480',  # オレンジ
            font=(FONT_NAME, 16, 'bold')
        )
        btn_select.pack(side='left', padx=10, ipadx=10, ipady=8)
        self.add_tooltip(btn_select, '応募バンドから出演バンドを選出します')

        btn_timetable = tk.Button(
            btn_frame_bottom,
            text='タイムテーブル作成\n※別ウィンドウ',
            width=14,
            height=2,
            command=self.make_timetable,
            bg="#d080ff",  # 紫
            font=(FONT_NAME, 16, 'bold')
        )
        btn_timetable.pack(side='left', padx=10, ipadx=10, ipady=8)
        self.add_tooltip(btn_timetable, 'タイムテーブル作成ウィンドウを開きます（別ウィンドウ）')
        
        # 設定ボタン（左下に固定配置）
        btn_settings = tk.Button(
            self.master,
            text='設定',
            width=10,
            height=1,
            command=self.show_settings,
            bg='#cccccc',  # グレー
            font=(FONT_NAME, 16, 'bold')
        )
        btn_settings.place(relx=0.0, rely=1.0, anchor='sw', x=20, y=-20) # 左下に固定配置
        # 設定ボタンにもツールチップ
        self.add_tooltip(btn_settings, 'アプリの各種設定を開きます')

        # 操作支援が有効な場合のみヘルプモードボタンを表示
        try:
            if self.settings.get('operation_support', False):
                btn_help_mode = tk.Button(
                    self.master,
                    text='ヘルプモード',
                    width=10,
                    height=1,
                    command=self.toggle_help_mode,
                    bg="#4375ff",
                    font=(FONT_NAME, 16, 'bold')
                )
                # 設定ボタンの右側に配置
                btn_help_mode.place(relx=0.0, rely=1.0, anchor='sw', x=180, y=-20)
                try:
                    self.add_tooltip(btn_help_mode, 'ヘルプモードに切り替えます（操作手順を案内します）')
                except Exception:
                    pass
        except Exception:
            pass

#上段のボタン
    #出欠をとる
    def show_attendance_date_select(self):
        """出席日付選択画面を表示"""
        self.clear()
        tk.Label(self.master, text='日付を選択', font=(FONT_NAME, 16, 'bold')).pack(pady=10)
        btn_top = tk.Button(self.master, text='トップに戻る', width=15, bg='#ff0000', font=(FONT_NAME, 12), command=self.show_top)
        btn_top.place(relx=1.0, rely=1.0, anchor='se', x=-20, y=-20)
        self.add_tooltip(btn_top, 'トップ画面に戻ります')
        btn_today = tk.Button(self.master, text='今日', width=15, bg='#66ff66', font=(FONT_NAME, 12), command=self.start_attendance_today)
        btn_today.pack(pady=5)
        self.add_tooltip(btn_today, '今日の日付で出欠登録を開始します')
        btn_other = tk.Button(self.master, text='別日', width=15, bg='#ff9900', font=(FONT_NAME, 12), command=self.start_attendance_otherday)
        btn_other.pack(pady=5)
        self.add_tooltip(btn_other, '別の日付で出欠登録を開始します')

    def start_attendance_today(self):
        today = datetime.datetime.now().strftime('%m/%d').lstrip('0').replace('/0', '/')
        self.start_attendance(date=today)

    def start_attendance_otherday(self):
        while True:
            date = simpledialog.askstring('日付入力', '日付を「M/D」形式で入力してください（例: 10/2）')
            if date is None:
                return  # キャンセル
            # M/DまたはMM/DD形式（0埋めなし/あり両対応）
            if re.fullmatch(r'\s*\d{1,2}/\d{1,2}\s*', date):
                # 月日が正しいかもチェック
                try:
                    m, d = map(int, date.strip().split('/'))
                    if 1 <= m <= 12 and 1 <= d <= 31:
                        self.start_attendance(date=date.strip())
                        return
                    else:
                        messagebox.showerror('入力エラー', '月日は正しい範囲で入力してください。')
                except Exception:
                    messagebox.showerror('入力エラー', '日付の形式が正しくありません。')
            else:
                messagebox.showerror('入力エラー', '日付は「M/D」形式で入力してください（例: 10/2）')

    def start_attendance(self, date):
        """出欠登録画面を表示"""
        # Excelから名簿データを取得
        self.df = pd.read_excel(FILE_PATH, sheet_name=SHEET_NAME, header=1, index_col=None)
        # 不要な「Unnamed: ...」列を除外
        self.df = self.df.loc[:, ~self.df.columns.str.contains('^Unnamed')]
        self.date = date
        # 日付列がなければ追加
        if date not in self.df.columns:
            self.df[date] = ''
        self.current_idx = 0
        self.show_attendance_entry()

    def show_attendance_entry(self):
        self.clear()
        if self.current_idx < 0:
            self.current_idx = 0
        if self.current_idx >= len(self.df):
            self.current_idx = len(self.df) - 1
        row = self.df.iloc[self.current_idx]
        def safe_str(val):
            import math
            return '' if val is None or (isinstance(val, float) and math.isnan(val)) else str(val)
        name = safe_str(row['氏名'])
        student_id = safe_str(row['学籍番号'])
        grade = safe_str(row['学年']) if '学年' in self.df.columns else ''
        faculty = safe_str(row['学部']) if '学部' in self.df.columns else ''

        info = f'No.{self.current_idx+1}  氏名: {name}\n学籍番号: {student_id}\n学年: {grade}  学部: {faculty}\n日付: {self.date}'
        tk.Label(self.master, text=info, font=(FONT_NAME, 14, 'bold'), justify='left').pack(pady=10)

        # 2段構成（3つ＋2つ）で出欠ボタンを配置
        mark_defs = [
            ('出席', '〇出席', '#66ff66'),
            ('連絡あり', '△連絡あり欠席', '#ffff66'),
            ('無断欠席', '×無断欠席', '#ff0000'),
            ('オ', 'オンライン', '#cccccc'),
            ('忌引', '忌引き等', '#cccccc'),
        ]
        btn_frame1 = tk.Frame(self.master)
        btn_frame1.pack(pady=3)
        btn_frame2 = tk.Frame(self.master)
        btn_frame2.pack(pady=3)
        btn_font = (FONT_NAME, 16)
        # 1段目（3つ）
        for mark, label, color in mark_defs[:3]:
            b = tk.Button(
                btn_frame1,
                text=label,
                width=12,
                height=1,
                bg=color,
                font=btn_font,
                command=lambda m=mark: self.set_attendance(m)
            )
            b.pack(side='left', padx=5)
            try:
                self.add_tooltip(b, f'{label} を記録します')
            except Exception:
                pass
        # 2段目（2つ）
        for mark, label, color in mark_defs[3:]:
            b2 = tk.Button(
                btn_frame2,
                text=label,
                width=12,
                height=1,
                bg=color,
                font=btn_font,
                command=lambda m=mark: self.set_attendance(m)
            )
            b2.pack(side='left', padx=5)
            try:
                self.add_tooltip(b2, f'{label} を記録します')
            except Exception:
                pass

        nav_frame = tk.Frame(self.master)
        nav_frame.pack(pady=10)
        btn_prev = tk.Button(nav_frame, text='前の人に戻る', bg='#ff9900', font=(FONT_NAME, 12), command=self.prev_person)
        btn_prev.pack(side='left', padx=10)
        self.add_tooltip(btn_prev, '前の人の出欠を表示します')
        btn_next_nav = tk.Button(nav_frame, text='次の人に進む', bg='#66ff66', font=(FONT_NAME, 12), command=self.next_person)
        btn_next_nav.pack(side='left', padx=10)
        self.add_tooltip(btn_next_nav, '次の人の出欠を表示します')

        btn_top = tk.Button(self.master, text='トップに戻る', width=15, bg='#ff0000', font=(FONT_NAME, 12), command=self.save_and_back_to_top)
        btn_top.place(relx=1.0, rely=1.0, anchor='se', x=-20, y=-20)
        self.add_tooltip(btn_top, '出欠を保存してトップに戻ります')

    def set_attendance(self, mark):
        self.df.at[self.current_idx, self.date] = mark
        self.next_person()

    def prev_person(self):
        self.current_idx -= 1
        if self.current_idx < 0:
            self.current_idx = 0
        self.show_attendance_entry()

    def next_person(self):
        self.current_idx += 1
        def is_empty_name(idx):
            import math
            if idx >= len(self.df):
                return True
            val = self.df.iloc[idx]['氏名']
            return (val is None) or (isinstance(val, float) and math.isnan(val)) or (str(val).strip() == '')
        if self.current_idx >= len(self.df) or is_empty_name(self.current_idx):
            messagebox.showinfo('完了', '全員分の出欠登録が完了しました。')
            self.save_and_back_to_top()
        else:
            self.show_attendance_entry()

    def save_and_back_to_top(self):
        # openpyxlで該当セルのみ書き換え
        try:
            wb = openpyxl.load_workbook(FILE_PATH)
            ws = wb[SHEET_NAME]
            # ① 2行目のG列(=7)以降で空白の列を探す（もし既に今日の日付の列があればそれを使う）
            target_col = None
            # まず今日の日付列があれば優先
            for col in range(1, ws.max_column + 1):
                if str(ws.cell(row=2, column=col).value) == str(self.date):
                    target_col = col
                    break
            if target_col is None:
                for col in range(7, ws.max_column + 1):
                    val = ws.cell(row=2, column=col).value
                    if val is None or str(val).strip() == '':
                        target_col = col
                        break
            # 見つからなければ末尾に追加
            if target_col is None:
                target_col = ws.max_column + 1
            # ヘッダが空なら日付を設定（左隣の書式をコピー）
            from copy import copy
            date_cell = ws.cell(row=2, column=target_col)
            if date_cell.value is None or str(date_cell.value).strip() == '':
                if target_col > 1:
                    left_cell = ws.cell(row=2, column=target_col-1)
                    date_cell.font = copy(left_cell.font)
                    date_cell.alignment = copy(left_cell.alignment)
                    date_cell.border = copy(left_cell.border)
                    date_cell.fill = copy(left_cell.fill)
                date_cell.value = self.date

            # 学籍番号列を探す（C/E列の位置が不明確なため、シート上の'学籍番号'というヘッダを探す）
            id_col = None
            for col in range(1, ws.max_column + 1):
                if str(ws.cell(row=2, column=col).value) == '学籍番号':
                    id_col = col
                    break
            if id_col is None:
                raise Exception('学籍番号列が見つかりません')

            # ② ③行目以降に出欠を書き込む（値のみ、書式は変更しない）
            for idx, row in self.df.iterrows():
                student_id = str(row['学籍番号'])
                excel_row = None
                for r in range(3, ws.max_row + 1):
                    if str(ws.cell(row=r, column=id_col).value) == student_id:
                        excel_row = r
                        break
                if excel_row is None:
                    continue
                cell = ws.cell(row=excel_row, column=target_col)
                # 左隣のセルの書式をコピー（なければ何もしない）
                from copy import copy
                if target_col > 1:
                    left_cell = ws.cell(row=excel_row, column=target_col-1)
                    cell.font = copy(left_cell.font)
                    cell.alignment = copy(left_cell.alignment)
                    cell.border = copy(left_cell.border)
                    cell.fill = copy(left_cell.fill)
                cell.value = row[self.date]
            wb.save(FILE_PATH)
            messagebox.showinfo('保存完了', 'Excelファイルを保存しました。')
        except Exception as e:
            messagebox.showerror('保存エラー', f'Excel保存に失敗しました: {e}')
        self.show_top()

    #出欠状況の確認
    def show_attendance_check(self):
        """出欠状況の確認画面"""
        self.clear()
        tk.Label(self.master, text='出欠状況の確認', font=(FONT_NAME, 16, 'bold')).pack(pady=10)
        # 期間選択説明
        info_label = tk.Label(self.master, text='計算期間を選択してください', font=(FONT_NAME, 12))
        info_label.pack(pady=10)
        # 開始日・終了日入力
        date_frame = tk.Frame(self.master)
        date_frame.pack(pady=10)
        # Excelから日付候補を取得
        date_candidates = self.get_available_dates()
        
        tk.Label(date_frame, text='開始日:', font=(FONT_NAME, 12)).pack(side='left', padx=5)
        start_combo = ttk.Combobox(date_frame, font=(FONT_NAME, 12), width=10, values=date_candidates)
        start_combo.pack(side='left', padx=5)
        
        tk.Label(date_frame, text='終了日:', font=(FONT_NAME, 12)).pack(side='left', padx=5)
        end_combo = ttk.Combobox(date_frame, font=(FONT_NAME, 12), width=10, values=date_candidates)
        end_combo.pack(side='left', padx=5)
        
        # 計算実行ボタン
        calc_btn = tk.Button(self.master, text='出席率を計算', width=15, height=2, bg='#80bfff', font=(FONT_NAME, 14, 'bold'), command=lambda: ac.calculate_attendance_rate(start_combo.get(), end_combo.get(), FILE_PATH, SHEET_NAME))
        calc_btn.pack(pady=20)
        try:
            self.add_tooltip(calc_btn, '指定した期間の出席率を計算して記録します')
        except Exception:
            pass
        
        # トップに戻るボタン（右下固定）
        back_btn = tk.Button(self.master, text='トップに戻る', width=15, height=1, bg='#ff0000', font=(FONT_NAME, 12), command=self.show_top)
        back_btn.place(relx=1.0, rely=1.0, anchor='se', x=-20, y=-20)
        try:
            self.add_tooltip(back_btn, 'トップ画面に戻ります')
        except Exception:
            pass

    def get_available_dates(self):
        """Excelファイルから利用可能な日付一覧を取得"""
        try:
            # 出欠状況シートを読み込み
            df = pd.read_excel(FILE_PATH, sheet_name=SHEET_NAME, header=1, index_col=None)
            
            # 日付列を特定（G列以降）
            date_list = []
            for col in df.columns[6:]:  # G列以降
                if pd.notna(col) and str(col).strip() and '/' in str(col):
                    date_str = str(col).strip()
                    # M/D形式かチェック
                    try:
                        parts = date_str.split('/')
                        if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
                            date_list.append(date_str)
                    except:
                        continue
            
            # 日付をソート（月/日の順）
            def date_sort_key(date_str):
                try:
                    month, day = map(int, date_str.split('/'))
                    return month * 100 + day
                except:
                    return 0
            
            date_list.sort(key=date_sort_key)
            return date_list
            
        except Exception as e:
            # エラーの場合は空のリストを返す
            return []

    #バンドを登録
    def register_band(self):
        """バンド登録画面を表示"""
        # 押下時に起動日（最終バンド登録日）を記録
        try:
            self.settings['last_startup'] = datetime.date.today().isoformat()
            try:
                self.save_settings()
            except Exception:
                pass
        except Exception:
            pass
        # ①～⑩日付割り当て画面
        def show_date_assign_dialog():
            from tkcalendar import Calendar
            assign_win = tk.Toplevel(self.master)
            assign_win.title('出演日割り当て')
            assign_win.geometry('320x600')
            tk.Label(assign_win, text='ライブの日程を設定してください', font=(FONT_NAME, 14, 'bold')).pack(pady=5)
            tk.Label(assign_win, text='ヒント：Googleフォームの出演可能日\nの選択肢と一致させる', font=(FONT_NAME, 12)).pack(pady=2)
            date_vars = {}
            label_vars = {}
            def open_calendar(num):
                cal_win = tk.Toplevel(assign_win)
                cal_win.title(f'[{num}]の日付選択')
                cal = Calendar(cal_win, selectmode='day', date_pattern='yyyy-mm-dd')
                cal.pack(padx=10, pady=10)
                def set_date():
                    date_vars[num].set(cal.get_date())
                    label_vars[num]['text'] = f'{num}日目: {cal.get_date()}'
                    cal_win.destroy()
                btn_cal_ok = tk.Button(cal_win, text='決定', font=(FONT_NAME, 12), command=set_date)
                btn_cal_ok.pack(pady=5)
                try:
                    self.add_tooltip(btn_cal_ok, 'この日付を確定します')
                except Exception:
                    pass
            frame = tk.Frame(assign_win)
            frame.pack(pady=5, fill='y')
            for i in range(1, 11):
                date_vars[i] = tk.StringVar(value='')
                label = tk.Label(frame, text=f'{i}日目: ', font=(FONT_NAME, 12), width=18, anchor='w', relief='groove')
                label.grid(row=i-1, column=0, padx=8, pady=4, sticky='w')
                label.bind('<Button-1>', lambda e, n=i: open_calendar(n))
                label_vars[i] = label
            def save_dates():
                self._date_assign_map = {}
                self.date_assignments = {}
                for i in range(1, 11):
                    self._date_assign_map[str(i)] = date_vars[i].get()
                    self.date_assignments[f'[{i}]'] = date_vars[i].get()
                assign_win.destroy()
                proceed_band_register()
            btn_assign_save = tk.Button(assign_win, text='保存', font=(FONT_NAME, 12, 'bold'), bg='#bfff80', width=12, command=save_dates)
            btn_assign_save.pack(pady=15)
            try:
                self.add_tooltip(btn_assign_save, '割り当てた日付を保存します')
            except Exception:
                pass
        # バンド登録本体
        def proceed_band_register():
            file_path = filedialog.askopenfilename(
                title='応募バンド情報のExcelファイルを選択',
                filetypes=[('Excelファイル', '*.xlsx;*.xls')]
            )
            if not file_path:
                return
            try:
                df_all = pd.read_excel(file_path)
                df_band = df_all.iloc[:, 2:]
            except Exception as e:
                messagebox.showerror('エラー', f'Excelファイルの読み込みに失敗しました:\n{e}')
                return
            self._band_register_queue = []
            for idx, row in df_band.iterrows():
                band_name = str(row.iloc[0])
                members_raw = str(row.iloc[2]) if len(row) > 2 else ''
                self._band_register_queue.append((band_name, members_raw, row))
            if self._band_register_queue:
                next_band = self._band_register_queue.pop(0)
                self.show_band_member_check(*next_band)
        # 割り当て済みならスキップ
        if not hasattr(self, '_date_assign_map'):
            show_date_assign_dialog()
        else:
            proceed_band_register()

    def _convert_perform_dates(self, perform_date_str):
        """
        バンドの出演希望日（例: '[0][2][5]'）の四角括弧数字を、ユーザーが割り当てた日付に置換する。
        [0]が含まれていた場合は、割り当て済みのすべての日付を;で連結して置換する。
        self.date_assignments: dict[str, str] で '[0]'～'[9]'→日付 の割り当てを保持している前提。
        未割り当ての場合はそのまま返す。
        """
        # ブラケット表記のみを抽出して対応する日付だけを返す
        # 例: '[1][2]テキスト' -> '2026-01-14;2026-01-15'（'テキスト' は削除）
        date_assignments = getattr(self, 'date_assignments', {})
        import re
        tokens = re.findall(r'\[\d+\]', str(perform_date_str))
        if not tokens:
            return ''
        out_dates = []
        for tok in tokens:
            if tok == '[0]':
                # 全ての割り当て済み日付を順に結合
                all_dates = [v for k, v in date_assignments.items() if v]
                out_dates.extend([d for d in all_dates if d])
            else:
                date = date_assignments.get(tok)
                if date:
                    out_dates.append(date)
        return ';'.join(out_dates)

    def show_band_member_check(self, band_name, members_raw, row_obj=None):
        """バンド登録のメンバー確認画面を表示"""
        # 既存のバンド登録画面を消去
        if hasattr(self, '_band_member_frame') and self._band_member_frame:
            self._band_member_frame.destroy()
        # トップ画面の他のウィジェットを消す
        self.clear()
        self._band_member_frame = tk.Frame(self.master)
        self._band_member_frame.pack(fill='both', expand=True)
        win = self._band_member_frame

        tk.Label(win, text=f'バンド名: {band_name}', font=(FONT_NAME, 14, 'bold')).pack(pady=5)

        # ③空白除去（改行はそのまま）
        members_str = re.sub(r'[ \u3000]', '', members_raw)
        member_lines = [line for line in members_str.splitlines() if line.strip()]

        # 名簿データ取得
        try:
            df_roster = pd.read_excel(globals().get('FILE_PATH', 'attend_data.xlsx'), sheet_name=globals().get('SHEET_NAME', '出欠状況'), header=1)
        except Exception as e:
            messagebox.showerror('エラー', f'名簿データの取得に失敗しました:\n{e}')
            win.destroy()
            return
        roster_names = list(df_roster['氏名'].dropna().astype(str))

        # ④一致度判定（30%以上）
        def similarity(a, b):
            # 単純な部分一致率（短い方の長さで）
            from difflib import SequenceMatcher
            return SequenceMatcher(None, a, b).ratio()

        matched_members = []
        for mline in member_lines:
            best_match = None
            best_score = 0
            for name in roster_names:
                score = similarity(mline, name)
                if score > best_score:
                    best_score = score
                    best_match = name
            if best_score >= 0.3:
                matched_members.append(best_match)

        # 横並びフレーム
        h_frame = tk.Frame(win)
        h_frame.pack(pady=0, side='top')

        # 左：読み込んだ情報
        left_frame = tk.Frame(h_frame)
        left_frame.pack(side='left', padx=10, fill='y')
        tk.Label(left_frame, text='読み込んだ情報', font=(FONT_NAME, 11)).pack()
        tk.Message(left_frame, text=members_raw, font=(FONT_NAME, 11), width=200).pack(pady=5)

        # 真ん中：自動判定メンバー
        center_frame = tk.Frame(h_frame)
        center_frame.pack(side='left', padx=10, fill='y')
        tk.Label(center_frame, text='自動判定メンバー', font=(FONT_NAME, 11, 'bold')).pack()
        for name in matched_members:
            tk.Label(center_frame, text=name, font=(FONT_NAME, 11)).pack(pady=0)

        # 右：⑥足りない分をコンボボックスで追加
        right_frame = tk.Frame(h_frame)
        right_frame.pack(side='left', padx=10, fill='y')
        max_members = 10
        remain = max_members - len(matched_members)
        add_vars = []
        if remain > 0:
            tk.Label(right_frame, text=f'追加メンバー（最大{remain}人まで選択）', font=(FONT_NAME, 11, 'bold')).pack()
            # 既に選ばれている人以外を候補に
            candidate_names = [n for n in roster_names if n not in matched_members]
            for i in range(remain):
                var = tk.StringVar()
                add_vars.append(var)
                cb = ttk.Combobox(right_frame, textvariable=var, values=candidate_names, font=(FONT_NAME, 11), width=20)
                cb.pack(pady=0)

        def save_band_to_excel():
            from openpyxl.utils import get_column_letter
            # メンバーリスト作成
            members_final = matched_members + [v.get() for v in add_vars if v.get()]
            while len(members_final) < 10:
                members_final.append('')
            # L:演奏時間
            l_val = str(row_obj.iloc[1]) if row_obj is not None and len(row_obj) > 1 else ''
            # M:出演可能日（必ず_convert_perform_datesで変換）
            m_val_raw = str(row_obj.iloc[3]) if row_obj is not None and len(row_obj) > 3 else ''
            # [n]のみ抽出し、日付に変換して;で連結
            bracket_nums = re.findall(r'\[\d+\]', m_val_raw)
            converted_dates = [self._convert_perform_dates(bn) for bn in bracket_nums]
            m_val = ';'.join([d for d in converted_dates if d])
            # N～P: opt1～opt3（ヘッダーに[opt1]～[opt3]含む列）
            n_val = o_val = p_val = ''
            if row_obj is not None:
                opt_cols = [i for i, col in enumerate(row_obj.index) if '[opt1]' in str(col)]
                n_val = str(row_obj.iloc[opt_cols[0]]) if len(opt_cols) > 0 else ''
                opt_cols2 = [i for i, col in enumerate(row_obj.index) if '[opt2]' in str(col)]
                o_val = str(row_obj.iloc[opt_cols2[0]]) if len(opt_cols2) > 0 else ''
                opt_cols3 = [i for i, col in enumerate(row_obj.index) if '[opt3]' in str(col)]
                p_val = str(row_obj.iloc[opt_cols3[0]]) if len(opt_cols3) > 0 else ''
            # 存在しない場合は空白（上記で空文字になるのでOK）
            # Q:その他（G列以降でopt1～3以外の最初の列）
            q_val = ''
            if row_obj is not None:
                for i, col in enumerate(row_obj.index[4:], 4):
                    if not any(f'[opt{n}]' in str(col) for n in range(1, 4)):
                        q_val = str(row_obj.iloc[i])
                        break
            band_row = [band_name]
            band_row.extend(members_final)
            band_row.extend([l_val, m_val, n_val, o_val, p_val, q_val])
            # 必要ならQ列まで空白で埋める
            while len(band_row) < 17:
                band_row.append('')
            # 18列目（R列）に0を追加
            band_row.append(0)
            # Excel保存
            try:
                wb = openpyxl.load_workbook(globals().get('FILE_PATH', 'attend_data.xlsx'))
                sheet_name = '登録済みバンド'
                if sheet_name not in wb.sheetnames:
                    ws = wb.create_sheet(sheet_name)
                else:
                    ws = wb[sheet_name]
                # A列が空の最初の行を探す
                row_idx = None
                for r in range(1, ws.max_row + 2):
                    if ws.cell(row=r, column=1).value in (None, ''):
                        row_idx = r
                        break
                if row_idx is None:
                    row_idx = ws.max_row + 1
                for col, val in enumerate(band_row, 1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.value = val
                    if ws.max_row >= 2:
                        from copy import copy
                        ref_cell = ws.cell(row=2, column=col)
                        try:
                            cell.font = copy(ref_cell.font)
                        except Exception:
                            pass
                wb.save(globals().get('FILE_PATH', 'attend_data.xlsx'))
                messagebox.showinfo('保存完了', 'バンド情報を保存しました。')
                win.destroy()
                # 次のバンドがあれば自動で次の登録画面へ
                if hasattr(self, '_band_register_queue') and self._band_register_queue:
                    next_band = self._band_register_queue.pop(0)
                    self.show_band_member_check(*next_band)
                else:
                    self.show_top()
            except Exception as e:
                messagebox.showerror('保存エラー', f'Excel保存に失敗しました:\n{e}')

        # --- 下部にバンド情報の詳細を表示・ボタン類配置 ---
        # 下部フレーム
        bottom_frame = tk.Frame(win)
        bottom_frame.pack(pady=10, side='bottom', fill='y')
        
        # ボタンフレーム
        btn_frame = tk.Frame(bottom_frame)
        btn_frame.pack(padx=25, pady=10, side='bottom')

        btn_save = tk.Button(btn_frame, text='登録', font=(FONT_NAME, 12, 'bold'), bg='#bfff80', width=10, height=1, command=save_band_to_excel)
        btn_save.pack(side='left', padx=0)
        self.add_tooltip(btn_save, 'バンド情報をExcelに保存します')

        def skip_to_next_band():
            win.destroy()
            if hasattr(self, '_band_register_queue') and self._band_register_queue:
                next_band = self._band_register_queue.pop(0)
                self.show_band_member_check(*next_band)
            else:
                self.show_top()

        btn_next = tk.Button(btn_frame, text='スキップ', font=(FONT_NAME, 12), bg='#ffe680', width=10, height=1, command=skip_to_next_band)
        btn_next.pack(side='left', padx=5)
        self.add_tooltip(btn_next, 'このバンドは登録せず次へ進みます')

        btn_clear = tk.Button(btn_frame, text='リセット', font=(FONT_NAME, 12), bg='#ff8080', width=10, height=1, command=self.clear_registered_bands)
        btn_clear.pack(side='left', padx=0)
        self.add_tooltip(btn_clear, '登録済みのバンド情報をすべて削除します')

        # バンド名の重複チェック
        duplicate_band = False
        try:
            wb = openpyxl.load_workbook(globals().get('FILE_PATH', 'attend_data.xlsx'))
            sheet_name = '登録済みバンド'
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for r in ws.iter_rows(min_row=1, max_col=1, values_only=True):
                    if r[0] == band_name:
                        duplicate_band = True
                        break
        except Exception:
            pass

        # 保存ボタンの上に警告表示
        if duplicate_band:
            tk.Label(bottom_frame, text='同じ名前のバンドがすでに登録されています', font=(FONT_NAME, 12, 'bold'), bg='#ffff66', fg='black').pack(pady=5)
        else:
            tk.Label(bottom_frame, text='バンド名の重複はありません', font=(FONT_NAME, 12, 'bold'), bg='#ccffcc', fg='black').pack(pady=5)
        
        info_frame = tk.Frame(win, relief='groove', bd=2)
        info_frame.pack(pady=1, side='bottom')
        # L列: 演奏時間
        l_val = str(row_obj.iloc[1]) if row_obj is not None and len(row_obj) > 1 else ''
        # M列: 出演可能日
        m_val = self._convert_perform_dates(str(row_obj.iloc[3]) if row_obj is not None and len(row_obj) > 3 else '')
        # N～P: opt1～opt3（ヘッダーに[opt1]～[opt3]含む列）
        n_val = o_val = p_val = ''
        if row_obj is not None:
            opt_cols = [i for i, col in enumerate(row_obj.index) if '[opt1]' in str(col)]
            n_val = str(row_obj.iloc[opt_cols[0]]) if len(opt_cols) > 0 else ''
            opt_cols2 = [i for i, col in enumerate(row_obj.index) if '[opt2]' in str(col)]
            o_val = str(row_obj.iloc[opt_cols2[0]]) if len(opt_cols2) > 0 else ''
            opt_cols3 = [i for i, col in enumerate(row_obj.index) if '[opt3]' in str(col)]
            p_val = str(row_obj.iloc[opt_cols3[0]]) if len(opt_cols3) > 0 else ''
        # Q:その他（G列以降でopt1～3以外の最初の列）
        q_val = ''
        if row_obj is not None:
            for i, col in enumerate(row_obj.index[4:], 4):
                if not any(f'[opt{n}]' in str(col) for n in range(1, 4)):
                    q_val = str(row_obj.iloc[i])
                    break
        # テキスト表示
        tk.Label(info_frame, text=f'演奏時間: {l_val}　　　出演可能日: {m_val}', font=(FONT_NAME, 11)).pack(anchor='w', pady=0)
        tk.Label(info_frame, text=f'オプション1: {n_val}', font=(FONT_NAME, 11)).pack(anchor='w', pady=0)
        tk.Label(info_frame, text=f'オプション2: {o_val}', font=(FONT_NAME, 11)).pack(anchor='w', pady=0)
        tk.Label(info_frame, text=f'オプション3: {p_val}', font=(FONT_NAME, 11)).pack(anchor='w', pady=0)
        tk.Label(info_frame, text=f'その他: {q_val}', font=(FONT_NAME, 11)).pack(anchor='w', pady=0)


        # トップに戻るボタン
        btn_top = tk.Button(win, text='トップに戻る', width=15, bg='#ff0000', font=(FONT_NAME, 12), command=self.show_top)
        btn_top.place(relx=1.0, rely=1.0, anchor='se', x=-20, y=-20)
        self.add_tooltip(btn_top, 'トップ画面に戻ります')

    def clear_registered_bands(self):
        """登録済みバンドのシートを空にする（確認ダイアログあり）。"""
        try:
            if not messagebox.askyesno('確認', '登録済みのバンド情報をすべて削除します。よろしいですか？\n\n※バンドを追加登録する場合は、「いいえ」を選択してください。', parent=self.master):
                return
        except Exception:
            # fallback if messagebox fails
            return
        try:
            wb = openpyxl.load_workbook(FILE_PATH)
            sheet_name = '登録済みバンド'
            if sheet_name not in wb.sheetnames:
                messagebox.showinfo('情報', '登録済みバンドのシートが見つかりません。', parent=self.master)
                return
            ws = wb[sheet_name]
            # ヘッダがある場合は1行目をヘッダとみなし2行目以降を消去
            first_cell = ws.cell(row=1, column=1).value
            start_row = 2 if first_cell and 'バンド' in str(first_cell) else 1
            for r in range(start_row, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).value = None
            wb.save(FILE_PATH)
            messagebox.showinfo('完了', '登録済みのバンド情報をすべて削除しました。', parent=self.master)
        except Exception as e:
            messagebox.showerror('エラー', f'バンド情報の削除に失敗しました:\n{e}', parent=self.master)

#下段のボタン
    #出演バンド選出
    def show_select_band(self):
        """出演バンド選出画面を表示"""
        self.clear()
        tk.Label(self.master, text='出演バンド選出', font=(FONT_NAME, 16, 'bold')).pack(pady=10)
        frm = tk.Frame(self.master)
        frm.pack(pady=10)
        # ①出席率の期間
        tk.Label(frm, text='出席率の期間', font=(FONT_NAME, 12)).grid(row=0, column=0, sticky='e', pady=4)
        period_list = []
        try:
            wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
            ws = wb['出席率記録']
            for col in range(7, ws.max_column + 1):
                val = ws.cell(row=2, column=col).value
                if val is not None and str(val).strip() != '':
                    period_list.append(str(val).strip())
        except Exception as e:
            period_list = []
        if not period_list:
            period_list = ['期間未登録']
        period_var = tk.StringVar()
        period_combo = ttk.Combobox(frm, textvariable=period_var, values=period_list, font=(FONT_NAME, 12), width=16, state='readonly')
        period_combo.grid(row=0, column=1, padx=8, pady=4, sticky='w')
        period_combo.current(0)
        # 出席率計算ボタン
        btn_calc = tk.Button(frm, text='出席率計算', font=(FONT_NAME, 11), bg='#80d4ff', command=self.show_attendance_check)
        btn_calc.grid(row=0, column=2, padx=8, pady=4, sticky='w')
        self.add_tooltip(btn_calc, '選択した期間の出席率を計算します')
        # ②募集枠数
        slots_var = tk.IntVar(value=8)
        tk.Label(frm, text='募集枠数', font=(FONT_NAME, 12)).grid(row=1, column=0, sticky='e', pady=4)
        tk.Entry(frm, textvariable=slots_var, font=(FONT_NAME, 12), width=10).grid(row=1, column=1, padx=8, pady=4)
        # ③募集総時間
        total_time_var = tk.IntVar(value=240)
        tk.Label(frm, text='募集総時間（分）', font=(FONT_NAME, 12)).grid(row=2, column=0, sticky='e', pady=4)
        tk.Entry(frm, textvariable=total_time_var, font=(FONT_NAME, 12), width=10).grid(row=2, column=1, padx=8, pady=4)
        # ④リハーサル・転換時間
        change_time_var = tk.IntVar(value=10)
        tk.Label(frm, text='リハーサル・転換時間（分）\n※1バンド当たり', font=(FONT_NAME, 12)).grid(row=3, column=0, sticky='e', pady=4)
        tk.Entry(frm, textvariable=change_time_var, font=(FONT_NAME, 12), width=10).grid(row=3, column=1, padx=8, pady=4)
        # 選出開始ボタン
        def on_select_band():
            period = period_var.get()
            slots = slots_var.get()
            total_time = total_time_var.get()
            change_time = change_time_var.get()
            self.select_band(period, slots, total_time, change_time)
        btn_start_select = tk.Button(frm, text='選出開始', font=(FONT_NAME, 14, 'bold'), bg='#bfff80', width=12, height=2, command=on_select_band)
        btn_start_select.grid(row=4, column=0, columnspan=2, pady=16)
        self.add_tooltip(btn_start_select, '指定条件で出演バンドを選出します')
        # トップに戻るボタン
        btn_top = tk.Button(self.master, text='トップに戻る', width=15, bg='#ff0000', font=(FONT_NAME, 12), command=self.show_top)
        btn_top.place(relx=1.0, rely=1.0, anchor='se', x=-20, y=-20)
        self.add_tooltip(btn_top, 'トップ画面に戻ります')

    def select_band(self, period, slots, total_time, change_time):
        result_text = bs.select_bands(period, slots, total_time, change_time, FILE_PATH, self.master)
        win = tk.Toplevel(self.master)
        win.title('出演バンド選出結果')
        txt = tk.Text(win, font=(FONT_NAME, 12), width=40, height=20)
        txt.pack(padx=10, pady=10)
        txt.insert('1.0', result_text)
        txt.config(state='normal')
        def select_all():
            txt.tag_add('sel', '1.0', 'end')
            self.master.clipboard_clear()
            self.master.clipboard_append(result_text)
        btn_copy = tk.Button(win, text='コピー', font=(FONT_NAME, 11), command=select_all)
        btn_copy.pack(pady=5)
        self.add_tooltip(btn_copy, '選出結果をクリップボードにコピーします')
        def show_applied_order():
            try:
                wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
                ws = wb['登録済みバンド']
                band_names = []
                for row in range(1, ws.max_row + 1):
                    name = ws.cell(row=row, column=1).value
                    r_val = ws.cell(row=row, column=18).value
                    if name and r_val == 1:
                        band_names.append(str(name))
                result = '\n'.join(band_names) if band_names else '該当バンドなし'
            except Exception as e:
                result = f'エラー: {e}'
            win2 = tk.Toplevel(win)
            win2.title('出演バンド（応募順）')
            txt2 = tk.Text(win2, font=(FONT_NAME, 12), width=40, height=20)
            txt2.pack(padx=10, pady=10)
            txt2.insert('1.0', result)
            txt2.config(state='normal')
            def select_all2():
                txt2.tag_add('sel', '1.0', 'end')
                self.master.clipboard_clear()
                self.master.clipboard_append(result)
            btn2 = tk.Button(win2, text='コピー', font=(FONT_NAME, 11), command=select_all2)
            btn2.pack(pady=5)
            try:
                self.add_tooltip(btn2, '一覧をコピーします')
            except Exception:
                pass
        btn_applied = tk.Button(win, text='応募順で表示', font=(FONT_NAME, 11), command=show_applied_order)
        btn_applied.pack(pady=5)
        self.add_tooltip(btn_applied, '応募順に出演バンドの一覧を表示します')

    #タイムテーブル作成
    def make_timetable(self):
        # 出演確定バンド情報をbands.csvに出力
        try:
            wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
            ws = wb['登録済みバンド']
            bands = []
            for row in range(1, ws.max_row + 1):
                r_val = ws.cell(row=row, column=18).value
                if r_val == 1:
                    band_name = ws.cell(row=row, column=1).value
                    play_time = ws.cell(row=row, column=12).value
                    perform_dates = ws.cell(row=row, column=13).value
                    opt1 = ws.cell(row=row, column=14).value or ''
                    opt2 = ws.cell(row=row, column=15).value or ''
                    opt3 = ws.cell(row=row, column=16).value or ''
                    other = ws.cell(row=row, column=17).value or ''
                    bands.append([str(band_name), str(play_time), str(perform_dates), str(opt1), str(opt2), str(opt3), str(other)])
            # CSV出力
            with open('bands.csv', 'w', encoding='utf-8', newline='') as f:
                import csv
                writer = csv.writer(f)
                writer.writerow(['バンド名', '演奏時間', '出演日', 'オプション1', 'オプション2', 'オプション3', 'その他'])
                for band in bands:
                    writer.writerow(band)
        except Exception as e:
            messagebox.showerror('CSV出力エラー', f'bands.csvの出力に失敗しました: {e}')
        # TopWindowを起動
        app = TopWindow()
        app.mainloop()

#設定画面
    def show_settings(self):
        # 設定一覧画面
        self.clear()
        tk.Label(self.master, text='設定', font=(FONT_NAME, 16, 'bold')).pack(pady=15)
        # 操作支援設定ボタン（Excelボタンに合わせたスタイル）
        btn_excel = tk.Button(
            self.master,
            text='Excelファイルの設定',
            font=(FONT_NAME, 14),
            fg='white',
            width=18,
            height=1,
            command=self.show_excel_file_settings,
            bg="#00bb44"
        )
        btn_excel.pack(pady=10)
        btn_op = tk.Button(
            self.master,
            text='操作支援機能',
            font=(FONT_NAME, 14),
            fg='white',
            width=18,
            height=1,
            command=self.show_operation_support_settings,
            bg="#35cbfd"
        )
        btn_op.pack(pady=10)
        try:
            self.add_tooltip(btn_excel, 'Excelファイル名の設定を開きます')
        except Exception:
            pass
        try:
            self.add_tooltip(btn_op, '操作支援機能の設定を開きます')
        except Exception:
            pass
        btn_top = tk.Button(self.master, text='トップに戻る', width=15, bg='#ff0000', font=(FONT_NAME, 12), command=self.show_top)
        btn_top.place(relx=1.0, rely=1.0, anchor='se', x=-20, y=-20)
        try:
            self.add_tooltip(btn_top, 'トップ画面に戻ります')
        except Exception:
            pass

    def show_excel_file_settings(self):
        # Excelファイル名の変更ダイアログ
        settings_win = tk.Toplevel(self.master)
        settings_win.title('Excelファイルの設定')
        settings_win.geometry('400x185')
        tk.Label(settings_win, text='Excelファイル名（パス）', font=(FONT_NAME, 12)).pack(pady=4)
        tk.Label(settings_win, text='※Excelファイルは同じフォルダ内に置いてください', font=(FONT_NAME, 10)).pack(pady=4)
        file_var = tk.StringVar(value=globals().get('FILE_PATH', 'attend_data.xlsx'))
        entry = tk.Entry(settings_win, textvariable=file_var, font=(FONT_NAME, 12), width=30)
        entry.pack(pady=5)

        def save_file_path():
            new_path = file_var.get().strip()
            if new_path:
                globals()['FILE_PATH'] = new_path
                messagebox.showinfo('設定', f'ファイル名を「{new_path}」に変更しました。')
                settings_win.destroy()
            else:
                messagebox.showerror('エラー', 'ファイル名を入力してください。')

        btn_save = tk.Button(settings_win, text='保存', font=(FONT_NAME, 12), command=save_file_path, width=10, height=1, bg='#bfff80')
        btn_save.pack(pady=15)
        try:
            self.add_tooltip(btn_save, 'Excelファイル名を保存します')
        except Exception:
            pass

    def show_operation_support_settings(self):
        """操作支援の設定ダイアログを表示（ボタンから呼ばれる）。"""
        settings_win = tk.Toplevel(self.master)
        settings_win.title('操作支援の設定')
        settings_win.geometry('420x180')
        tk.Label(settings_win, text='操作支援機能の設定', font=(FONT_NAME, 14, 'bold')).pack(pady=8)
        op_var = tk.BooleanVar(value=self.settings.get('operation_support', True))
        chk = tk.Checkbutton(settings_win, text='操作支援機能を有効にする', variable=op_var, font=(FONT_NAME, 12))
        chk.pack(pady=8)

        def save_op_setting():
            self.settings['operation_support'] = bool(op_var.get())
            try:
                self.save_settings()
                messagebox.showinfo('設定', '操作支援の設定を保存しました。', parent=self.master)
                settings_win.destroy()
            except Exception as e:
                messagebox.showerror('エラー', f'設定の保存に失敗しました:\n{e}', parent=self.master)

        btn_frame_ops = tk.Frame(settings_win)
        btn_frame_ops.pack(pady=10)

        btn_save = tk.Button(btn_frame_ops, text='保存', font=(FONT_NAME, 12), command=save_op_setting, width=12, height=1, bg='#bfff80')
        btn_save.pack(side='left', padx=6)
        try:
            self.add_tooltip(btn_save, '操作支援の設定を保存します')
        except Exception:
            pass

        def rerun_walkthrough():
            # ウォークスルー既読フラグをリセットして即時表示
            self.settings['seen_walkthrough'] = False
            try:
                self.save_settings()
            except Exception:
                pass
            try:
                messagebox.showinfo('操作案内', 'ウォークスルーを再表示します。', parent=self.master)
            except Exception:
                pass
            try:
                self.show_walkthrough()
            except Exception as e:
                try:
                    messagebox.showerror('エラー', f'ウォークスルーの表示に失敗しました:\n{e}', parent=self.master)
                except Exception:
                    pass

        btn_rerun = tk.Button(btn_frame_ops, text='操作案内を再表示', font=(FONT_NAME, 12), command=rerun_walkthrough, width=16, height=1, bg='#ffd480')
        btn_rerun.pack(side='left', padx=6)
        try:
            self.add_tooltip(btn_rerun, 'ウォークスルーを再表示します')
        except Exception:
            pass

#共通処理
    def __init__(self, master):
        self.master = master
        master.title('ロック部 出席管理')
        # ウィンドウの×ボタンに確認ダイアログを設定
        try:
            self.master.protocol("WM_DELETE_WINDOW", self.on_close)
        except Exception:
            pass
        # 設定読み込み（操作支援など）
        self.load_settings()
        # 前回起動日が設定されている場合、30日以上経過していれば確認ダイアログを表示
        try:
            prev = self.settings.get('last_startup')
            today = datetime.date.today()
            if prev:
                try:
                    prev_date = datetime.date.fromisoformat(prev)
                    delta_days = (today - prev_date).days
                    if delta_days >= 30:
                        try:
                            msg = f'最後のバンド登録から{delta_days}日経過しています。新たにバンドを募集しますか？\n「はい」を選ぶと登録済みのバンド情報を削除します。'
                            if messagebox.askyesno('お久しぶりです！', msg, parent=self.master):
                                try:
                                    self.clear_registered_bands()
                                except Exception:
                                    pass
                        except Exception:
                            pass
                except Exception:
                    pass
        except Exception:
            pass
        self.show_top()
        # 起動時ウォークスルー表示（操作支援が有効かつ未表示の場合）
        try:
            self.maybe_show_walkthrough()
        except Exception:
            pass

    def clear(self):
        for widget in self.master.winfo_children():
            widget.destroy()

    def on_close(self):
        if messagebox.askokcancel('確認', 'アプリを終了しますか？', parent=self.master):
            try:
                self.master.destroy()
            except Exception:
                pass


#操作支援機能
    def add_tooltip(self, widget, text, delay=300):
        """ウィジェットにツールチップを追加（操作支援が有効な場合のみ表示）。"""
        def on_enter(event):
            try:
                if not self.settings.get('operation_support', False):
                    return
            except Exception:
                return
            try:
                widget._tooltip_after = widget.after(delay, lambda: self._show_tooltip(widget, text))
            except Exception:
                pass

        def on_leave(event):
            try:
                if hasattr(widget, '_tooltip_after'):
                    widget.after_cancel(widget._tooltip_after)
                    del widget._tooltip_after
            except Exception:
                pass
            self._hide_tooltip(widget)

        widget.bind('<Enter>', on_enter)
        widget.bind('<Leave>', on_leave)
        widget.bind('<ButtonPress>', lambda e: self._hide_tooltip(widget))

    def _show_tooltip(self, widget, text):
        try:
            if hasattr(widget, '_tooltip') and widget._tooltip:
                return
            x = widget.winfo_rootx() + 20
            y = widget.winfo_rooty() + widget.winfo_height() + 5
            tw = tk.Toplevel(self.master)
            tw.wm_overrideredirect(True)
            tw.wm_geometry(f"+{x}+{y}")
            lbl = tk.Label(tw, text=text, font=(FONT_NAME, 10), bg='#ffffe0', justify='left', relief='solid', bd=1)
            lbl.pack(ipadx=4, ipady=2)
            widget._tooltip = tw
        except Exception:
            pass

    def _hide_tooltip(self, widget):
        try:
            if hasattr(widget, '_tooltip') and widget._tooltip:
                widget._tooltip.destroy()
                widget._tooltip = None
        except Exception:
            pass

    def get_config_path(self, filename='settings.json'):
        # 実行パスを考慮して設定ファイルパスを返す
        if hasattr(sys, '_MEIPASS'):
            base_dir = os.path.dirname(sys.executable)
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(base_dir, filename)

    def load_settings(self):
        path = self.get_config_path()
        self.settings = {}
        try:
            if os.path.exists(path):
                with open(path, 'r', encoding='utf-8') as f:
                    import json
                    self.settings = json.load(f)
        except Exception:
            self.settings = {}
        # デフォルトは操作支援を有効にする
        self.settings.setdefault('operation_support', True)
        # ウォークスルー既読フラグのデフォルト
        self.settings.setdefault('seen_walkthrough', False)

    def save_settings(self):
        path = self.get_config_path()
        try:
            with open(path, 'w', encoding='utf-8') as f:
                import json
                json.dump(self.settings, f, ensure_ascii=False, indent=2)
        except Exception as e:
            messagebox.showerror('エラー', f'設定の保存に失敗しました:\n{e}', parent=self.master)

    def maybe_show_walkthrough(self):
        try:
            if self.settings.get('operation_support', False) and not self.settings.get('seen_walkthrough', False):
                self.show_walkthrough()
        except Exception:
            pass

    def show_walkthrough(self):
        """モーダル形式のウォークスルーを表示（短い手順）。"""
        steps = [
            ('ようこそ', '次の画面から操作手順を説明します。\n順番に実行することですべての機能を利用できます。'),
            ('①出欠をとる', 'トップの「出欠をとる」から日付を選択し、出欠を記録します。'),
            ('②出席率の計算', '「出欠状況の確認」で開始日と終了日を選び、出席率を計算・記録します。'),
            ('③バンド登録', '「バンドを登録」で応募データを読み込み、メンバー確認のうえExcelに登録します。'),
            ('④バンド選出', '「出演バンド選出」で条件を指定して出演バンドを選出します。'),
            ('⑤タイムテーブル作成', '「タイムテーブル作成」を開き、出演順を編集・保存・出力します（別ウィンドウ）。'),
            ('設定と引き継ぎ', '「設定」で操作支援のON/OFFを切り替え、「他のPCへ引き継ぎ」でデータをエクスポートできます。')
        ]

        win = tk.Toplevel(self.master)
        win.title('操作手順案内')
        win.geometry('520x220')
        win.transient(self.master)
        win.grab_set()

        idx_var = tk.IntVar(value=0)
        text_title = tk.Label(win, text=steps[0][0], font=(FONT_NAME, 14, 'bold'))
        text_title.pack(pady=(12,4))
        text_body = tk.Label(win, text=steps[0][1], font=(FONT_NAME, 12), wraplength=480, justify='left')
        text_body.pack(padx=12)

        chk_var = tk.BooleanVar(value=False)
        chk = tk.Checkbutton(win, text='今後表示しない', variable=chk_var, font=(FONT_NAME, 11))
        chk.pack(pady=6)

        btn_frame = tk.Frame(win)
        btn_frame.pack(pady=6)

        def update_step():
            i = idx_var.get()
            text_title.config(text=steps[i][0])
            text_body.config(text=steps[i][1])
            if i == 0:
                btn_back.config(state='disabled')
            else:
                btn_back.config(state='normal')
            if i == len(steps)-1:
                btn_next.config(text='完了', bg="#00ff62" )
            else:
                btn_next.config(text='次へ', bg="#00FF62")

        def on_next():
            i = idx_var.get()
            if i < len(steps)-1:
                idx_var.set(i+1)
                update_step()
            else:
                # 完了
                if chk_var.get():
                    self.settings['seen_walkthrough'] = True
                    try:
                        self.save_settings()
                    except Exception:
                        pass
                else:
                    # デフォルトでは完了で既読にする
                    self.settings['seen_walkthrough'] = True
                    try:
                        self.save_settings()
                    except Exception:
                        pass
                win.grab_release()
                win.destroy()

        def on_back():
            i = idx_var.get()
            if i > 0:
                idx_var.set(i-1)
                update_step()

        def on_close():
            # 閉じたときは「今後表示しない」が有効なら既読にする
            if chk_var.get():
                self.settings['seen_walkthrough'] = True
                try:
                    self.save_settings()
                except Exception:
                    pass
            win.grab_release()
            win.destroy()

        btn_back = tk.Button(btn_frame, text='戻る', width=10, command=on_back)
        btn_back.pack(side='left', padx=6)
        btn_next = tk.Button(btn_frame, text='次へ', width=10, command=on_next, bg='#00ff62')
        btn_next.pack(side='left', padx=6)
        btn_close = tk.Button(btn_frame, text='閉じる', width=10, command=on_close, bg='#ff0000')
        btn_close.pack(side='left', padx=6)

        update_step()
        win.protocol('WM_DELETE_WINDOW', on_close)

    def toggle_help_mode(self):
        """ヘルプモードをトグルして設定を保存する。"""
        try:
            current = bool(self.settings.get('help_mode', False))
            self.settings['help_mode'] = not current
            try:
                self.save_settings()
                self.show_top()
            except Exception:
                pass
        except Exception:
            pass


    def not_implemented(self):
        messagebox.showinfo('未実装', 'この機能はまだ実装されていません。')

if __name__ == '__main__':
    root = tk.Tk()
    root.geometry('700x600')  # ウィンドウサイズを大きく設定
    root.minsize(700, 600)    # ウィンドウの最小サイズを設定
    default_font = tkfont.nametofont("TkDefaultFont")
    default_font.configure(family="Yu Gothic UI", size=10)
    root.option_add("*Font", default_font)
    # 既存コードで FONT_NAME を参照しているフォント指定があるため、
    # 名称 FONT_NAME を FONT_NAME にマッピングする named font を作成します。
    try:
        tkfont.Font(name=FONT_NAME, family=FONT_NAME, size=10)
    except Exception:
        pass
    app = AttendanceApp(root)
    root.mainloop()