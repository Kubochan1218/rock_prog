import os, sys, json, shutil, openpyxl
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import messagebox
import customtkinter as ctk
import config

class TimetableView(ctk.CTkFrame):
    def get_json_path(self, filename, mode="r"):
        if hasattr(sys, '_MEIPASS'):
            # PyInstallerでexe化した場合も、常にexeのある場所を参照
            base_dir = os.path.dirname(sys.executable)
        else:
            # スクリプト実行時はtop.pyのあるディレクトリ
            base_dir = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(base_dir, "..", filename)

    def __init__(self, master, app, **kwargs):
        super().__init__(master, fg_color="transparent", **kwargs)
        self.app = app
        LIVE_JSON_PATH = self.app.get_config_path('live_info.json')
        self.existing_lives = {}
        self.schedules = []        
        self.tabs = {}
        self.used_band_names = set()
        self.tab_control = None
        self.order_area = None
        self.top_widgets = []  # トップ画面ウィジェットを管理
        self.filter_options = {'オプション1': '', 'オプション2': '', 'オプション3': ''}  # フィルタ条件
        self._last_font_copied = False
        
        # 既存のJSONデータが存在すれば読み込む
        if os.path.exists(LIVE_JSON_PATH):
            try:
                with open(LIVE_JSON_PATH, 'r', encoding='utf-8') as f:
                    self.existing_lives = json.load(f)
            except Exception:
                pass

        self.load_band_infos()
        self.create_widgets()

    def _ensure_font_file(self, filename, friendly_name="フォント"):
        if hasattr(sys, '_MEIPASS'):
            # PyInstallerでexe化した場合も、常にexeのある場所を参照
            app_dir = os.path.dirname(sys.executable)
        else:
            # スクリプト実行時はtop.pyのあるディレクトリ
            app_dir = os.path.dirname(os.path.abspath(__file__))
        target_path = os.path.join(app_dir, "..", filename)
        if os.path.exists(target_path):
            self._last_font_copied = False
            return target_path

        windir = os.environ.get("WINDIR", "C:\\Windows")
        fonts_dir = os.path.join(windir, "Fonts")
        candidates = []
        try:
            for f in os.listdir(fonts_dir):
                if f.lower() == filename.lower() or filename.split('.')[0].lower() in f.lower():
                    candidates.append(os.path.join(fonts_dir, f))
        except Exception:
            candidates = []

        if not candidates:
            messagebox.showerror(f"{friendly_name} コピー失敗", f"{filename} が見つかりません。\nシステムフォントフォルダに {filename} が存在しないようです。", parent=self)
            return None

        for src in candidates:
            try:
                shutil.copy2(src, target_path)
                self._last_font_copied = True
                return target_path
            except Exception:
                continue

        messagebox.showerror(f"{friendly_name} コピー失敗", f"{filename} のコピーに失敗しました。手動で {filename} をアプリフォルダに置いてください。", parent=self)
        return None

    def back_to_main_menu(self):
        if messagebox.askyesno("確認", "タイムテーブル作成画面を閉じてメインメニューに戻りますか？\n保存していない変更は破棄されます。", parent=self):
            self.app.show_top()

    def create_widgets(self):
        title_label = ctk.CTkLabel(self, text="タイムテーブル作成", font=config.FONT_TITLE)
        title_label.pack(pady=15, anchor="w")
        
        select_live_frame = ctk.CTkFrame(self, fg_color="transparent")
        select_live_frame.pack(pady=10, fill='x', padx=10)

        def on_live_select(event):
            choice = self.live_name_combo.get()
            self.schedules.clear()
            for sch in self.existing_lives[choice].get('schedules', []):
                date_str = sch.get('date', '')
                start_str = sch.get('start', '')
                end_str = sch.get('end', '')
                saved_bands = sch.get('bands', [])
                try:
                    date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
                    self.schedules.append({'date': date_obj, 'start': start_str, 'end': end_str, 'saved_bands': saved_bands})
                except ValueError:
                    continue
            self.show_order_area()

        ctk.CTkLabel(select_live_frame, text='ライブ名を選択:', font=config.FONT_LABEL_BUTTON).pack(side='left', padx=0)
        live_names_list = list(self.existing_lives.keys())
        self.live_name_combo = ctk.CTkComboBox(
            select_live_frame, 
            values=live_names_list if live_names_list else [""], 
            font=(config.FONT_NAME, 16), 
            width=240,
            state="readonly",
            command=on_live_select
        )
        self.live_name_combo.set("") # 初期値は空
        self.live_name_combo.pack(side='left', padx=10)

    def show_order_area(self):
        if not self.schedules:
            messagebox.showerror("日程エラー", "「ライブ管理」からライブの日程を1日以上追加してください。", parent=self)
            return
        # トップ画面を非表示にする
        for widget in self.top_widgets:
            widget.pack_forget()
            
        if self.order_area:
            self.order_area.destroy()
        self.order_area = ctk.CTkFrame(self, fg_color="transparent")
        self.order_area.pack(fill="both", expand=True)
        self.create_order_area()

    def create_order_area(self):
        menu_frame = ctk.CTkFrame(self.order_area, fg_color="transparent")
        menu_frame.pack(fill="x", padx=10, pady=3)

        btn_export_excel = ctk.CTkButton(menu_frame, text="Excel出力", command=self.export_excel, font=config.FONT_LABEL_BUTTON, width=160, fg_color="#07ca6f", text_color="white", hover_color="#05964f")
        btn_export_excel.pack(side="left", padx=10)
        
        btn_option = ctk.CTkButton(menu_frame, text="絞り込み", command=self.show_option_dialog, font=config.FONT_LABEL_BUTTON, width=80, fg_color="#fff9c4", text_color="black", hover_color="#fff59d")
        btn_option.pack(side="left", padx=10)

        # Tabviewへのアップグレード
        self.tab_control = ctk.CTkTabview(self.order_area)
        self.tab_control.pack(expand=True, fill="both", padx=10, pady=5)
        
        self.tabs = {}
        self.used_band_names = set()
        self.create_tabs()

    def show_option_dialog(self):
        win = ctk.CTkToplevel(self.winfo_toplevel())
        win.title("絞り込みオプション")
        win.geometry("380x180")
        win.after(200, lambda: win.focus())
        win.attributes("-topmost", True)
        win.grab_set()  # モーダル化
        
        labels = ["オプション1", "オプション2", "オプション3"]
        entries = {}
        for i, opt in enumerate(labels):
            option_frame = ctk.CTkFrame(win, fg_color="transparent")
            option_frame.pack(pady=5, fill='x')
            ctk.CTkLabel(option_frame, text=opt, width=120, font=config.FONT_LABEL_BUTTON).pack(side="left", anchor="w", padx=10)
            ent = ctk.CTkEntry(option_frame, width=180, font=config.FONT_LABEL_BUTTON)
            ent.insert(0, self.filter_options.get(opt, ''))
            ent.pack(side="left", anchor="w", padx=10)
            entries[opt] = ent
            
        def on_ok():
            for opt in labels:
                self.filter_options[opt] = entries[opt].get().strip()
            win.destroy()
            for tab in self.tabs:
                self.refresh_combo(tab)
                
        def on_clear():
            for opt in labels:
                self.filter_options[opt] = ''
                entries[opt].delete(0, tk.END)
            win.destroy()
            for tab in self.tabs:
                self.refresh_combo(tab)
                
        button_frame = ctk.CTkFrame(win, fg_color="transparent")
        button_frame.pack(side="bottom", anchor="center", pady=10)
        btn_ok = ctk.CTkButton(button_frame, text="OK", command=on_ok, font=config.FONT_LABEL_BUTTON, width=100, fg_color="#c8e6c9", text_color="black", hover_color="#a5d6a7")
        btn_ok.pack(side="left", padx=5)
        btn_clear = ctk.CTkButton(button_frame, text="全解除", command=on_clear, font=config.FONT_LABEL_BUTTON, width=100, fg_color="#ffcdd2", text_color="black", hover_color="#ef9a9a")
        btn_clear.pack(side="left", padx=5)

    def load_band_infos(self):
        """Excelファイルからバンド情報を読み込む"""
        wb = openpyxl.load_workbook(config.FILE_PATH, data_only=True)
        ws = wb['登録済みバンド']
        self.bands = []
        for row in range(1, ws.max_row + 1):
            r_val = ws.cell(row=row, column=18).value
            if r_val == 1:
                band_name = ws.cell(row=row, column=1).value
                play_time = ws.cell(row=row, column=12).value
                perform_dates = ws.cell(row=row, column=13).value
                opt1 = ws.cell(row=row, column=14).value or ''
                opt2 = ws.cell(row=row, column=15).value or ''
                opt3 = ws.cell(row=row, column=16).value or ''
                other = ws.cell(row=row, column=17).value or ''
                live_name = ws.cell(row=row, column=19).value or ''
                self.bands.append({"band_name" : str(band_name), "play_time" : str(play_time), "perform_dates" : str(perform_dates), "opt1" : str(opt1), "opt2" : str(opt2), "opt3" : str(opt3), "other" : str(other), "live_name" : str(live_name)})
    
    def create_tabs(self):
        for sched in self.schedules:
            label_text = sched['date'].strftime("%m/%d")
            self.tab_control.add(label_text)
            tab = self.tab_control.tab(label_text)
            
            top_frame = ctk.CTkFrame(tab, fg_color="transparent")
            top_frame.pack(fill="x", padx=10, pady=5)
            
            label = ctk.CTkLabel(top_frame, text="バンドを選択：", font=config.FONT_LABEL_BUTTON)
            label.pack(side="left")
            
            combo = ctk.CTkComboBox(top_frame, width=200, font=config.FONT_LABEL_BUTTON, state="readonly")
            combo.pack(side="left", padx=5)
            
            btn_add = ctk.CTkButton(top_frame, text="追加", command=lambda c=combo, t=label_text: self.add_band(c, t), font=config.FONT_LABEL_BUTTON, width=80, fg_color="#e0f7fa", text_color="black", hover_color="#b2dfdb")
            btn_add.pack(side="left", padx=5)
            
            btn_special = ctk.CTkButton(top_frame, text="特別枠追加", command=lambda t=label_text: self.add_special_frame(t), font=config.FONT_LABEL_BUTTON, width=100, fg_color="#ffe0b2", text_color="black", hover_color="#ffcc80")
            btn_special.pack(side="left", padx=5)
            
            label_info = ctk.CTkLabel(top_frame, text="バンド名をクリックして\nバンド情報を表示", font=config.FONT_LABEL_BUTTON, justify="left")
            label_info.pack(side="left", padx=5)
            
            # Canvasエリア一式を ctk.CTkScrollableFrame の1行に置き換え
            frame_container = ctk.CTkScrollableFrame(tab)
            frame_container.pack(fill="both", expand=True, padx=5, pady=5)
            
            available_bands = [b for b in self.bands if str(sched['date']) in b['perform_dates'] and b['live_name'] == self.live_name_combo.get()]
            combo.configure(values=[b['band_name'] for b in available_bands])
            if available_bands:
                combo.set(available_bands[0]['band_name'])
            else:
                combo.set("")

            bands_list = []
            for bname in sched.get('saved_bands', []):
                # 登録されているバンド情報から演奏時間(分)を探して取得
                band_obj = self.find_band_by_name(bname, available_bands)
                if band_obj:
                    minutes = band_obj['play_time']
                else:
                    minutes = 30  # 万が一バンドデータから見つからなかった場合のデフォルト値
                    
                # 画面描画用の形式に変換してリストに詰める
                bands_list.append({'type': 'band', 'name': bname, 'minutes': minutes})
                # 出演確定リストに登録（他の日程の選択肢から除外するため）
                self.used_band_names.add(bname)

            self.tabs[label_text] = {
                "combo": combo,
                "frame": frame_container,
                "bands": bands_list,
                "band_objs": available_bands,
                "tab_label": label_text
            }

            self.update_band_frames(label_text)
        
        for t in self.tabs.keys():
            self.refresh_combo(t)

    def add_band(self, combo, tab):
        """選択されたバンドを指定されたタブに追加する"""
        # combo: CTkComboBoxのインスタンス, tab: タブのラベル（例: "10/12"）
        band_name = combo.get()
        if not band_name:
            return
        tab_info = self.tabs[tab]
        if any(b['type']=='band' and b['name']==band_name for b in tab_info["bands"]):
            messagebox.showwarning("重複追加", f"{band_name} はすでに他の日程で追加されています。", parent=self)
            return
        band = self.find_band_by_name(band_name, tab_info["band_objs"])
        if not band:
            messagebox.showerror("エラー", "バンド情報が見つかりません", parent=self)
            return
        tab_info["bands"].append({'type': 'band', 'name': band_name, 'minutes': band['play_time']})
        self.used_band_names.add(band_name)
        for t in self.tabs.keys():
            self.refresh_combo(t)
        self.update_band_frames(tab)
        self.auto_save_live_bands()

    def add_special_frame(self, tab):
        def on_ok():
            kind = var_kind.get()
            try:
                minutes = int(entry_min.get())
            except ValueError:
                messagebox.showerror("エラー", "分数は整数で入力してください", parent=self)
                return
            if kind == 'リハ':
                band_name = combo_band.get().strip()
                if not band_name:
                    messagebox.showerror("エラー", "バンド名を選択してください", parent=self)
                    return
                item = {'type': 'rehearsal', 'name': band_name, 'minutes': minutes}
            elif kind == '休憩':
                item = {'type': 'break', 'minutes': minutes}
            elif kind == '転換':
                item = {'type': 'change', 'minutes': minutes}
            else:
                return
            self.tabs[tab]["bands"].append(item)
            win.destroy()
            self.update_band_frames(tab)
            self.auto_save_live_bands()

        win = ctk.CTkToplevel(self.winfo_toplevel())
        win.title("特別枠追加")
        win.geometry("360x200")
        win.after(200, lambda: win.focus())

        win.attributes("-topmost", True)
        win.grab_set()  # モーダル化
        
        kinds_frame = ctk.CTkFrame(win, fg_color="transparent")
        kinds_frame.pack(pady=5, fill='x')
        ctk.CTkLabel(kinds_frame, text="種別", font=config.FONT_LABEL_BUTTON).pack(side="left", anchor="w", padx=10, pady=5)
        var_kind = tk.StringVar(value='休憩')
        kinds = ['休憩', '転換', 'リハ']
        combo_kind = ctk.CTkComboBox(kinds_frame, values=kinds, variable=var_kind, state="readonly", width=180, font=config.FONT_LABEL_BUTTON)
        combo_kind.pack(side="right", padx=10)
        
        minutes_frame = ctk.CTkFrame(win, fg_color="transparent")
        minutes_frame.pack(pady=5, fill='x')
        ctk.CTkLabel(minutes_frame, text="分数", font=config.FONT_LABEL_BUTTON).pack(side="left", anchor="w", padx=10, pady=5)
        entry_min = ctk.CTkEntry(minutes_frame, width=180, justify='center', font=config.FONT_LABEL_BUTTON)
        entry_min.pack(side="right", padx=10)
        
        band_frame = ctk.CTkFrame(win, fg_color="transparent")
        band_frame.pack(pady=5, fill='x')
        ctk.CTkLabel(band_frame, text="バンド名(リハのみ)", font=config.FONT_LABEL_BUTTON).pack(side="left", anchor="w", padx=10, pady=5)
        all_band_names = [b['band_name'] for b in self.tabs[tab]['band_objs']]
        combo_band = ctk.CTkComboBox(band_frame, values=all_band_names, width=180, font=config.FONT_LABEL_BUTTON)
        combo_band.pack(side="right", padx=10)
        
        btn = ctk.CTkButton(win, text="OK", command=on_ok, font=config.FONT_LABEL_BUTTON)
        btn.pack(side="bottom", padx=10, pady=10)
        
        def on_kind_change(choice):
            if choice == 'リハ':
                combo_band.configure(state='normal')
            else:
                combo_band.set("")
                combo_band.configure(state='disabled')
                
        combo_kind.configure(command=on_kind_change)
        on_kind_change(var_kind.get())

    def update_band_frames(self, tab):
        tab_info = self.tabs[tab]
        frame = tab_info["frame"]
        for widget in frame.winfo_children():
            widget.destroy()
        sched = None
        tab_label = tab_info.get("tab_label")
        for s in self.schedules:
            if s['date'].strftime("%m/%d") == tab_label:
                sched = s
                break
        if not sched:
            return
        end_time_limit = datetime.strptime(sched['end'], "%H:%M")
        current_time = datetime.strptime(sched['start'], "%H:%M")
        
        for idx, item in enumerate(tab_info["bands"]):
            if item['type'] == 'band':
                name = item['name']
                minutes = item['minutes']
                label_text = f"{name}（{minutes}分）"
                band_info = self.find_band_by_name(name, self.bands)
                def show_band_info(event, band_info=band_info):
                    if band_info:
                        info = f"バンド名: {band_info['band_name']}\n演奏時間: {band_info['play_time']}分\n出演日: {band_info['perform_dates']}\nオプション1: {band_info['opt1']}\nオプション2: {band_info['opt2']}\nオプション3: {band_info['opt3']}\nその他: {band_info['other']}"
                        if hasattr(band_info, 'options'):
                            for k, v in band_info['options'].items():
                                info += f"\n{k}: {v}"
                        if hasattr(band_info, 'other') and band_info['other']:
                            info += f"\nその他: {band_info['other']}"
                        messagebox.showinfo("バンド情報", info, parent=self)
            elif item['type'] == 'break':
                name = None
                minutes = item['minutes']
                label_text = f"休憩（{minutes}分）"
            elif item['type'] == 'change':
                name = None
                minutes = item['minutes']
                label_text = f"転換（{minutes}分）"
            elif item['type'] == 'rehearsal':
                name = item['name']
                minutes = item['minutes']
                label_text = f"リハ（{name}）（{minutes}分）"
            else:
                continue
                
            start_str = current_time.strftime("%H:%M")
            end_time = current_time + timedelta(minutes=int(minutes))
            end_str = end_time.strftime("%H:%M")
            
            bg_color = None
            if item['type'] == 'band' and band_info is not None:
                other_val = getattr(band_info, 'other', None)
                if other_val is not None and str(other_val).lower() != 'nan' and str(other_val).strip() != '':
                    bg_color = '#b2ebf2'  # 水色
            if bg_color is None and end_time > end_time_limit:
                bg_color = '#fff59d'  # 黄色
                
            # ctk.CTkFrame でのコンテナ
            if bg_color:
                wrapper = ctk.CTkFrame(frame, corner_radius=4, border_width=1, border_color="gray", fg_color=bg_color, height=40)
            else:
                wrapper = ctk.CTkFrame(frame, corner_radius=4, border_width=1, border_color="gray", height=40)
                
            wrapper.pack(fill="x", expand=True, pady=2, padx=5)
            wrapper.pack_propagate(False)
            
            label = ctk.CTkLabel(wrapper, text=f"{start_str}～{end_str} {label_text}", anchor="w", font=config.FONT_LABEL_BUTTON, text_color="black" if bg_color else None)
            label.pack(side="left", fill="x", expand=True, padx=10)
            
            if item['type'] == 'band':
                label.bind('<Button-1>', show_band_info)
                
            btn_del = ctk.CTkButton(wrapper, text="×", width=30, height=25, fg_color="transparent", text_color="red", hover_color="lightcoral", font=config.FONT_LABEL_BUTTON)
            btn_del.pack(side="right", padx=5)
            
            btn_down = ctk.CTkButton(wrapper, text="↓", width=25, height=25, font=config.FONT_LABEL_BUTTON)
            btn_down.pack(side="right", padx=2)
            
            btn_up = ctk.CTkButton(wrapper, text="↑", width=25, height=25, font=config.FONT_LABEL_BUTTON)
            btn_up.pack(side="right", padx=2)
            
            def remove_item(idx=idx):
                removed = tab_info["bands"][idx]
                del tab_info["bands"][idx]
                if removed.get("type") == "band":
                    self.used_band_names.discard(removed["name"])
                    for t in self.tabs:
                        self.refresh_combo(t)
                self.update_band_frames(tab)
                self.auto_save_live_bands()
                
            btn_up.configure(command=lambda idx=idx: self.move_band(tab, idx, -1))
            btn_down.configure(command=lambda idx=idx: self.move_band(tab, idx, 1))
            btn_del.configure(command=remove_item)
            
            if end_time > end_time_limit:
                warn = ctk.CTkLabel(wrapper, text="時刻超過", fg_color='#fff59d', text_color="black", font=config.FONT_LABEL_BUTTON)
                warn.pack(side="right", padx=10)
            current_time = end_time

    def move_band(self, tab, idx, direction):
        tab_info = self.tabs[tab]
        new_idx = idx + direction
        if 0 <= new_idx < len(tab_info["bands"]):
            tab_info["bands"][idx], tab_info["bands"][new_idx] = tab_info["bands"][new_idx], tab_info["bands"][idx]
            self.update_band_frames(tab)
            self.auto_save_live_bands()

    def auto_save_live_bands(self):
        """バンドの出演順が変わったときに、live_info.jsonへ自動保存する関数"""
        # 現在コンボボックスで選択されているライブ名を取得
        live_key = self.live_name_combo.get()
        if not live_key or live_key not in self.existing_lives:
            return  # ライブが選択されていない場合は何もしない

        # 最新の live_info.json を一度読み込む（他データの誤消去を防ぐため）
        json_path = self.get_json_path("live_info.json")
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                live_data = json.load(f)
        except Exception:
            live_data = self.existing_lives

        if live_key not in live_data:
            return

        # ライブの日程リストを取得
        schedules = live_data[live_key].get('schedules', [])

        # 各日程ごとにタイムテーブルの上から順にバンド名を抽出して格納
        for sch in schedules:
            sch_date_str = sch.get('date', '')  # 例: "2026-06-20"
            target_bands = []

            # 現在表示中の全スケジュール（self.schedules）から日付が一致するタブを探す
            for s in self.schedules:
                if s['date'].strftime("%Y-%m-%d") == sch_date_str:
                    tab_label = s['date'].strftime("%m/%d")  # タブのキー（例: "06/20"）
                    
                    if tab_label in self.tabs:
                        # タイムテーブルに並んでいる全枠を取得
                        band_items = self.tabs[tab_label]["bands"]
                        # 純粋な「バンド枠」かつ「名前があるもの」だけを順番に抽出
                        target_bands = [
                            item['name'] for item in band_items 
                            if item.get('type') == 'band' and item.get('name')
                        ]
                    break

            # JSONデータ構造内の対象日程に "bands" リストを追加・更新
            sch['bands'] = target_bands

        # live_info.json に上書き保存
        try:
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(live_data, f, ensure_ascii=False, indent=4)
            # メモリ上のデータも最新状態に更新
            self.existing_lives = live_data
        except Exception as e:
            messagebox.showerror("自動保存エラー", f"JSONの書き込みに失敗しました: {e}", parent=self)

    def refresh_combo(self, tab):
        tab_info = self.tabs[tab]
        used_names = self.used_band_names
        def match_option(band):
            i = 0
            for opt, val in self.filter_options.items():
                i += 1
                v = val.strip()
                if v == '':
                    continue
                band_opt = band[f"opt{i}"].strip()
                if band_opt != v:
                    return False
            return True
        filtered = [b["band_name"] for b in tab_info["band_objs"] if b["band_name"] not in used_names and match_option(b)]
        if all(v.strip() == '' for v in self.filter_options.values()):
            filtered = [b["band_name"] for b in tab_info["band_objs"] if b["band_name"] not in used_names]
            
        tab_info["combo"].configure(values=filtered)
        if filtered:
            tab_info["combo"].set(filtered[0])
        else:
            tab_info["combo"].set("")

    def find_band_by_name(self, name, band_list):
        for b in band_list:
            if b["band_name"] == name:
                return b
        return None

    def export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
        except ImportError:
            messagebox.showerror("Excel出力エラー", "openpyxlライブラリが必要です。\npip install openpyxl でインストールしてください。", parent=self)
            return
        font_file = os.path.join(os.path.dirname(__file__), "meiryo.ttc")
        if not os.path.exists(font_file):
            if not self._ensure_font_file("meiryo.ttc", "Meiryo"):
                return

        wb = openpyxl.Workbook()
        for tab, tabinfo in self.tabs.items():
            tab_title = tabinfo.get("tab_label")
            safe_title = tab_title.replace("/", "-")
            ws = wb.create_sheet(title=safe_title)
            ws.append(["開始時刻", "～", "終了時刻", "枠の名前"])
            header_font = Font(name="Meiryo", size=12, bold=True)
            for col in range(1, 5):
                ws.cell(row=ws.max_row, column=col).font = header_font
            for idx, item in enumerate(tabinfo["bands"]):
                sched = None
                for s in self.schedules:
                    if s['date'].strftime("%m/%d") == tabinfo.get("tab_label"):
                        sched = s
                        break
                if not sched:
                    continue
                current_time = datetime.strptime(sched['start'], "%H:%M")
                for i in range(idx):
                    current_time += timedelta(minutes=int(tabinfo["bands"][i]["minutes"]))
                start_str = current_time.strftime("%H:%M")
                end_time = current_time + timedelta(minutes=int(item["minutes"]))
                end_str = end_time.strftime("%H:%M")
                name = item.get("name") if item.get("type") != "break" and item.get("type") != "change" else None
                if item["type"] == "band":
                    row = [start_str, "～", end_str, name]
                    ws.append(row)
                    for col in range(1, 5):
                        ws.cell(row=ws.max_row, column=col).fill = PatternFill(fill_type=None)
                        ws.cell(row=ws.max_row, column=col).font = Font(name="Meiryo", size=11, color="000000")
                else:
                    if item["type"] == "break":
                        disp_name = f"休憩（{item['minutes']}分）"
                    elif item["type"] == "change":
                        disp_name = f"転換（{item['minutes']}分）"
                    elif item["type"] == "rehearsal":
                        disp_name = f"リハ（{item['name']}）（{item['minutes']}分）"
                    else:
                        disp_name = ""
                    row = [start_str, "～", end_str, disp_name]
                    ws.append(row)
                    for col in range(1, 5):
                        ws.cell(row=ws.max_row, column=col).fill = PatternFill(start_color="888888", end_color="888888", fill_type="solid")
                        ws.cell(row=ws.max_row, column=col).font = Font(name="Meiryo", size=11, color="FFFFFF")
                for col in range(1, 5):
                    ws.cell(row=ws.max_row, column=col).alignment = Alignment(horizontal="left", vertical="center")
            ws.column_dimensions["A"].width = 10
            ws.column_dimensions["B"].width = 4
            ws.column_dimensions["C"].width = 10
            ws.column_dimensions["D"].width = 24
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        try:
            wb.save(f"【{self.live_name_combo.get()}】タイムテーブル.xlsx")
            msg = f"【{self.live_name_combo.get()}】タイムテーブル.xlsx を出力しました。"
            if getattr(self, '_last_font_copied', False):
                msg += "\n※フォントファイルコピー済み"
            messagebox.showinfo("Excel出力", msg, parent=self)
            self._last_font_copied = False
        except Exception as e:
            messagebox.showerror("Excel出力エラー", str(e), parent=self)
