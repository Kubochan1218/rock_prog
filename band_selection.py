# バンド選出ロジック関連の関数・クラスをここに分割して記述
# 引数：period, slots, total_time, change_time, file_path

import openpyxl
from tkinter import messagebox

# 応募順で表示ボタン
def select_bands(period, slots, total_time, change_time, file_path, master):
    try:
        slots = int(slots)
        total_time = int(total_time)
        change_time = int(change_time)
    except Exception:
        return [], '', '枠数・総時間・リハ転換時間は整数で入力してください。'
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws_rate = wb['出席率記録']
        ws_band = wb['登録済みバンド']
    except Exception as e:
        return [], '', f'Excelファイルの読み込みに失敗しました\n{e}'
    # 期間列特定
    period_col = None
    for col in range(7, ws_rate.max_column + 1):
        val = ws_rate.cell(row=2, column=col).value
        if str(val).strip() == period:
            period_col = col
            break
    if period_col is None:
        return [], '', f'選択した期間「{period}」の出席率データがありません。'
    # 氏名列特定
    name_col = None
    for col in range(1, ws_rate.max_column + 1):
        val = ws_rate.cell(row=2, column=col).value
        if str(val).strip() == '氏名':
            name_col = col
            break
    if name_col is None:
        return [], '', '出席率記録シートに「氏名」列がありません。'
    # 個人ポイント初期化
    member_points = {}
    member_rates = {}
    for row in range(3, ws_rate.max_row + 1):
        name = ws_rate.cell(row=row, column=name_col).value
        rate = ws_rate.cell(row=row, column=period_col).value
        if name is None or str(name).strip() == '':
            continue
        try:
            rate_val = float(rate)
        except:
            rate_val = 0
        member_rates[name] = rate_val
        if rate_val >= 100:
            pt = 6
        elif rate_val >= 80:
            pt = 5
        elif rate_val >= 60:
            pt = 4
        elif rate_val >= 40:
            pt = 3
        elif rate_val >= 20:
            pt = 2
        elif rate_val > 0:
            pt = 1
        else:
            pt = 0
        member_points[name] = pt
    # バンド情報取得
    band_list = []
    for row in range(1, ws_band.max_row + 1):
        band_name = ws_band.cell(row=row, column=1).value
        if not band_name:
            continue
        members = [ws_band.cell(row=row, column=col).value for col in range(2, 12)]
        members = [m for m in members if m and str(m).strip() != '']
        try:
            play_time = int(ws_band.cell(row=row, column=12).value)
        except:
            play_time = 0
        band_list.append({
            'name': band_name,
            'members': members,
            'play_time': play_time,
            'row': row
        })
        ws_band.cell(row=row, column=18).value = 0  # R列クリア
    # 選出ループ＋ログ
    selected_bands = []
    used_time = 0
    log_lines = []
    log_lines.append(f'=== バンド選出ログ ===')
    log_lines.append(f'期間: {period}  枠数: {slots}  総時間: {total_time}  リハ・転換: {change_time}')
    log_lines.append('')
    round_num = 1
    while len(selected_bands) < slots and band_list:
        log_lines.append(f'--- 選出ラウンド {round_num} ---')
        for band in band_list:
            pts = [member_points.get(m, 0) for m in band['members']]
            band['point'] = sum(pts) / len(pts) if pts else 0
            log_lines.append(f'バンド: {band["name"]}  メンバー: {", ".join(band["members"])}  ポイント: {band["point"]:.2f}  演奏時間: {band["play_time"]}')
        max_point = max(b['point'] for b in band_list)
        max_bands = [b for b in band_list if b['point'] == max_point]
        log_lines.append(f'最大ポイント: {max_point:.2f}  選出バンド: {", ".join(b["name"] for b in max_bands)}')
        add_time = sum(b['play_time'] + change_time for b in max_bands)
        if used_time + add_time > total_time:
            # 時間超過しても最大ポイントのバンドは全て出演確定
            for b in max_bands:
                t = b['play_time'] + change_time
                selected_bands.append(b)
                used_time += t
                log_lines.append(f'出演確定: {b["name"]} (累計時間: {used_time})')
            log_lines.append(f'時間制限を超過しましたが最大ポイントバンド全て出演確定')
            break
        else:
            for b in max_bands:
                selected_bands.append(b)
                used_time += b['play_time'] + change_time
                log_lines.append(f'出演確定: {b["name"]} (累計時間: {used_time})')
        member_decrement = {}
        for b in max_bands:
            for m in b['members']:
                member_decrement[m] = member_decrement.get(m, 0) + 1
        for m, dec in member_decrement.items():
            before = member_points.get(m, 0)
            member_points[m] = before - dec
            log_lines.append(f'  {m}: {before}→{member_points[m]} (出演確定数: {dec})')
        band_list = [b for b in band_list if b not in max_bands]
        round_num += 1
        log_lines.append('')
    # ログ保存
    try:
        with open('select_band_log.txt', 'w', encoding='utf-8') as f:
            f.write('\n'.join(log_lines))
    except Exception as e:
            messagebox.showerror('エラー', f'ログ保存に失敗しました: {e}')
    # 出演確定バンドのR列（18列目）に1を書き込む
    try:
        for b in selected_bands:
            ws_band.cell(row=b['row'], column=18).value = 1
        wb.save(file_path)
    except Exception as e:
        messagebox.showerror('エラー', f'出演確定バンドのR列書き込みに失敗しました: {e}')

    # 結果表示（バンド名のみ）
    result_text = ''
    for i, b in enumerate(selected_bands, 1):
        result_text += f'{b["name"]}\n'
    if not selected_bands:
        result_text += '条件に合致するバンドがありませんでした。\n'
    return result_text
